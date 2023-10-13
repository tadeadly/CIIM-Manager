import pandas as pd
from openpyxl import load_workbook
import os


def write_data_to_excel(src_path, target_date, target_directory, mappings, start_row=4):
    target_datetime = pd.to_datetime(target_date, format="%d/%m/%y", errors="coerce")
    formatted_target_date = target_datetime.strftime("%d.%m.%y")
    report_filename = f"CIIM Report Table {formatted_target_date}.xlsx"
    target_report_path = os.path.join(target_directory, report_filename)

    # Using mappings to determine columns to load from the source
    usecols_value = list(mappings.values())
    df = pd.read_excel(src_path, skiprows=1, usecols=usecols_value)

    # Filter data
    target_df = filter_by_date(df, "Date [DD/MM/YY]", target_datetime)

    # Open the target workbook
    target_workbook = load_workbook(filename=target_report_path)
    target_worksheet = target_workbook.active

    # Write data
    for row_idx, (index, row_data) in enumerate(target_df.iterrows(), start=start_row):
        for col_idx, header in enumerate(mappings.keys(), 2):  # Starting from column B
            src_header = mappings[header]
            target_worksheet.cell(
                row=row_idx, column=col_idx, value=row_data[src_header]
            )

    # Format Date column
    date_col_idx = list(mappings.keys()).index("Date [DD/MM/YY]") + 2
    format_datetime_column(
        target_worksheet, date_col_idx, start_row, target_worksheet.max_row
    )

    # Format Observations column (assuming "Observations" is always a key in your mappings)
    observations_col_idx = list(mappings.keys()).index("Observations") + 2
    format_observations_column(
        target_worksheet, observations_col_idx, start_row, target_worksheet.max_row
    )

    target_workbook.save(target_report_path)
    print(f"Report for {formatted_target_date} has been updated and saved.")


def write_data_to_report(src_path, target_date, target_directory, mappings):
    write_data_to_excel(src_path, target_date, target_directory, mappings)


def write_data_to_previous_report(src_path, target_date, target_directory, mappings):
    # Prompt the user for the starting row
    start_row_delay = simpledialog.askinteger(
        "Input", "Enter the starting row:", minvalue=4
    )
    start_row_delay = int(start_row_delay)
    # If the user cancels the prompt or doesn't enter a valid number, exit the function
    if not start_row_delay:
        return

    write_data_to_excel(
        src_path, target_date, target_directory, mappings, start_row=start_row_delay
    )


def save_to_excel():
    global delay_excel_workbook

    if not selected_tl:
        return

    full_file_path = delays_folder_path / f"{selected_tl}.xlsx"
    delay_excel_workbook = load_workbook(filename=full_file_path)
    delay_excel_worksheet = delay_excel_workbook["Sheet1"]

    # Direct assignments using ENTRIES_CONFIG
    for entry_name, config in ENTRIES_CONFIG.items():
        cell_address = config["cell"]
        entry = globals()[entry_name]
        delay_excel_worksheet[cell_address] = entry.get()

    # Update cell H8 with the latest username
    latest_username = get_latest_username_from_file()
    if latest_username:
        delay_excel_worksheet["H8"] = latest_username

    print(f"Active username : {latest_username}")
    delay_excel_workbook.save(str(full_file_path))
    clear_cells()
    load_from_excel()
    line_status()
    print(f"Saved successfully : {selected_tl}")


def status_check():
    global status_color

    if (
        start_time == 1
        and end_time == 1
        and reason_var == 1
        and worker1_var == 1
        and vehicle1_var == 1
    ):
        set_config(frame3_status, text="Completed", foreground="green")

        status_color = 1
    else:
        set_config(frame3_status, text="Not completed", foreground="#E83845")
        status_color = 0


def set_entry_status(entry, var_name, default_val=0):
    if entry.get() == "":
        entry.config(style="danger.TEntry")
        globals()[var_name] = default_val
    else:
        entry.config(style="success.TEntry")
        globals()[var_name] = 1
