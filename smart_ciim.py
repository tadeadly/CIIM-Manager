import os
import re
import shutil
import time
from datetime import timedelta, datetime
from pathlib import Path
from tkinter import *
from tkinter import filedialog, simpledialog, messagebox, Menu
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import ImageTk, Image
from ctypes import windll
import pyglet
from ttkbootstrap.dialogs import Querybox
import ttkbootstrap as ttk
from ttkbootstrap.tooltip import ToolTip


def center_window(window, parent):
    window.update_idletasks()  # To ensure the size is calculated

    # Calculate position x, y coordinates
    x = parent.winfo_x() + (parent.winfo_width() // 2) - (window.winfo_width() // 2)
    y = parent.winfo_y() + (parent.winfo_height() // 2) - (window.winfo_height() // 2)

    window.geometry(f"+{x}+{y}")
    window.deiconify()  # Show the window


def define_related_paths():
    """Define all paths relative to the global CIIM_FOLDER_PATH."""
    base_path = CIIM_DIR_PATH

    paths = {
        "delays": base_path / "General Updates" / "Delays+Cancelled works",
        "faults": base_path / "General Updates" / "Fault Report Management" / "Electrification Control Center Fault "
                                                                              "Report Management 2.0.xlsx",
        "passdown": base_path / "Pass Down",
        "templates": base_path / "Important doc" / "Empty reports (templates)",
        "procedure": base_path / "Important Doc" / "Protocols" / "CIIM procedure test2.0.xlsx",
    }

    return paths


def get_ciim_dir_path_from_file(file_path):
    """Retrieve the CIIM folder path from the given file path."""
    return file_path.parent.parent.parent


def select_const_wp():
    """
    Opens a file dialog for the user to select an Excel file.
    """
    global construction_wp_var, construction_wp_path
    pattern = "WW*Construction Work Plan*.xlsx"
    path = filedialog.askopenfilename(filetypes=[("Excel Files", pattern)])

    if path:  # Check if a path was actually selected
        construction_wp_path = Path(path)
        construction_wp_var.set(construction_wp_path.name)  # Set the StringVar to just the filename
        # After path has been set, update the dates
        update_dates_based_on_file()
    # If no path was selected, simply do nothing (i.e., leave the entry as is)

    return construction_wp_path if path else None


def open_const_wp(event=None):
    """
    Handle the opening and reading of the construction work plan file.
    Fetches paths for the Construction Plan and CIIM folder, and extracts unique dates from the worksheet.
    At the end, prompts the user to provide their username.
    """
    global construction_wp_path, CIIM_DIR_PATH, cp_dates, username

    construction_wp_path = select_const_wp()
    if not construction_wp_path:
        return

    construction_wp_workbook = load_workbook(filename=construction_wp_path)
    print(f"The Construction Plan Path is : {construction_wp_path}")

    CIIM_DIR_PATH = get_ciim_dir_path_from_file(construction_wp_path)
    print(f"The CIIM folder Path is : {CIIM_DIR_PATH}")

    cp_dates = extract_unique_dates_from_worksheet(
        construction_wp_workbook["Const. Plan"]
    )
    print(f"Dates : {cp_dates}")
    construction_wp_workbook.close()

    username = username_var.get()

    return construction_wp_path, CIIM_DIR_PATH


def update_dates_based_on_file():
    """
    Update the unique dates based on the selected construction work plan file.
    """
    global construction_wp_path, CIIM_DIR_PATH, cp_dates

    if not construction_wp_path or construction_wp_path == Path("/"):
        return

    construction_wp_workbook = load_workbook(filename=construction_wp_path)
    CIIM_DIR_PATH = get_ciim_dir_path_from_file(construction_wp_path)
    cp_dates = extract_unique_dates_from_worksheet(
        construction_wp_workbook["Const. Plan"]
    )
    construction_wp_workbook.close()

    # Update any other widgets or global variables that depend on these dates here
    # For example, if you have a Listbox displaying the dates, you'd update it here.


def extract_unique_dates_from_worksheet(worksheet):
    """
    Extract unique dates from a given worksheet column.

    Returns:
        list: List of unique dates extracted from the worksheet, sorted in ascending order.
    """
    unique_dates = set()
    for cell in worksheet["D"]:
        date_value_str = process_date_cell(cell)
        if date_value_str:
            unique_dates.add(date_value_str)

    return sorted(list(unique_dates))


def process_date_cell(cell):
    """
    Processes a given cell's value to extract the date.
    Handles both datetime objects and strings representing dates.

    Returns:
        str: String representation of the date in the format YYYY-MM-DD.
             Returns None if no valid date is found.
    """
    if not cell.value:
        return None

    if isinstance(cell.value, datetime):
        return cell.value.date().isoformat()

    try:
        date_value = datetime.strptime(cell.value, "%d/%m/%Y").date()
        return date_value.isoformat()
    except ValueError:
        return None


def get_filtered_team_leaders(construction_wp_worksheet, date):
    """
    Extracts team leader names and their corresponding indexes from a worksheet for a specific date.
    Excludes team leaders that match a predefined blacklist.

    Returns:
        tuple: Contains a list of team leader names and a list of their corresponding indexes in the worksheet.
    """
    global TL_BLACKLIST, tl_index

    maxrow = construction_wp_worksheet.max_row
    team_leaders_list, tl_index = [], []

    for index in range(3, maxrow):
        cell_obj = construction_wp_worksheet.cell(row=index, column=4)
        if pd.Timestamp(cell_obj.value) == date:
            tl_cell_value = construction_wp_worksheet.cell(row=index, column=13).value
            if tl_cell_value:
                tl_name = re.sub("[-0123456789)(.]", "", str(tl_cell_value)).strip()
                if tl_name not in TL_BLACKLIST:
                    team_leaders_list.append(tl_name)
                    tl_index.append(index)

    return team_leaders_list, tl_index


def dc_combo_selected(event):
    """
    Handle date selection from the dates_combobox and update relevant variables.

    Note:
        Also reads the construction worksheet to get the relevant list of team leaders for the selected date.
    """

    global dc_year, dc_month, dc_week, dc_day, tl_index, dc_selected_date

    dc_selected_date = pd.Timestamp(dates_combobox.get())
    dc_day, dc_month, dc_year = [
        dc_selected_date.strftime(pattern) for pattern in ["%d", "%m", "%Y"]
    ]
    dc_week = dc_selected_date.strftime("%U")

    construction_wp_workbook = load_workbook(filename=construction_wp_path)
    construction_wp_worksheet = construction_wp_workbook["Const. Plan"]

    team_leaders_list, tl_index = get_filtered_team_leaders(
        construction_wp_worksheet, dc_selected_date
    )

    dc_tl_listbox.delete(0, END)
    for tl_name in team_leaders_list:
        dc_tl_listbox.insert(END, tl_name)

    construction_wp_workbook.close()

    # noinspection PyArgumentList
    dates_combobox.configure(bootstyle="default")
    dc_create_button.config(state=NORMAL)
    dc_create_all_button.config(state=NORMAL)


def update_combo_list():
    """
    Update the values of dates_combobox and dm_dates_combobox with cp_dates values.
    """

    dates_combobox["values"] = cp_dates
    dm_dates_combobox["values"] = cp_dates


def initialize_progress_bar_window(title, max_value):
    """
    Initialize and show a progress bar window.
    """
    progress_win = ttk.Toplevel()
    progress_win.withdraw()
    center_window(progress_win, progress_win.master)
    progress_win.title(title)
    progress_label = ttk.Label(progress_win, text="Creating reports...")
    progress_label.grid(row=0, column=0, padx=20)
    progress_bar = ttk.Progressbar(progress_win, orient="horizontal", length=300, mode="determinate")
    progress_bar.grid(row=1, column=0, padx=20, pady=20)
    progress_bar["maximum"] = max_value
    progress_bar["value"] = 0
    progress_win.update()

    return progress_win, progress_bar, progress_label


def create_all_delays():
    """
    Handle the creation of delay reports for all team leaders in the list box.
    """
    global dc_selected_team_leader, tl_num

    if not dc_selected_date:
        messagebox.showerror("Error", "Select the date and try again")
        return

    confirm = messagebox.askokcancel(
        "Confirmation",
        "Are you sure you want to create delay reports for ALL Team leaders?"
    )
    if not confirm:
        return

    # Get all the names from the listbox
    all_names = dc_tl_listbox.get(0, END)

    progress_win, progress_bar, progress_label = initialize_progress_bar_window(
        "Creating Delay Reports for All Team Leaders...", len(all_names))

    successful_creations = 0  # Counter to accumulate successful creations

    # Loop through all the names
    for index, name in enumerate(all_names):
        dc_selected_team_leader = name
        tl_num = tl_index[index]  # Assuming tl_index matches the order in the listbox
        success, status_msg = create_delay_wb()
        if success:
            dc_tl_listbox.itemconfig(index, background="#ED969D")
            successful_creations += 1

        # Update progress bar and label
        progress_bar["value"] += 1
        progress_label.config(text=f"Creating report {progress_bar['value']} of {progress_bar['maximum']}...")
        progress_win.update()

    # Close the progress bar window
    progress_win.destroy()

    # Consolidated message
    if successful_creations > 0:
        consolidated_msg = f"{successful_creations} Delay Reports were created successfully!"
        messagebox.showinfo("Success", consolidated_msg)
    else:
        messagebox.showerror("Error", "No reports were created!")


def dc_on_listbox_create():
    """
    Handle the event of a double click on the list box of team leaders.
    """
    global dc_selected_team_leader, tl_num

    if not dc_selected_date:
        messagebox.showerror("Error", "Select the date and try again")
        return

    dc_listbox_selection_indices = dc_tl_listbox.curselection()

    progress_win, progress_bar, progress_label = initialize_progress_bar_window("Creating Delay Reports...",
                                                                                len(dc_listbox_selection_indices))

    successful_creations = 0  # Counter to accumulate successful creations

    # Loop through the tuple of selected indices
    for index in dc_listbox_selection_indices:
        dc_selected_team_leader = str(dc_tl_listbox.get(index))
        tl_num = tl_index[index]

        success, status_msg = create_delay_wb()
        if success:
            dc_tl_listbox.itemconfig(index, background="#ED969D")
            successful_creations += 1

        # Update progress bar and label
        progress_bar["value"] += 1
        progress_label.config(text=f"Creating report {progress_bar['value']} of {progress_bar['maximum']}...")
        progress_win.update()

    # Close the progress bar window
    progress_win.destroy()

    # Consolidated message
    if successful_creations > 0:
        consolidated_msg = "Delay Reports were created successfully!"
        messagebox.showinfo("Success", consolidated_msg)
    else:
        messagebox.showerror("Error", "No reports were created!")


def create_delay_wb():
    """
    Define paths for delays and create a delay report for the selected date and team leader.

    Note:
        Copies the delay report template, populates it, and saves it to the appropriate path.
    """

    paths = define_related_paths()
    delays_path = paths["delays"]

    dc_year_path = delays_path / dc_year
    dc_week_path = dc_year_path / f"WW{dc_week}"
    dc_day_dir = f"{dc_day}.{dc_month}.{dc_year[2:]}"
    dc_day_path = dc_week_path / dc_day_dir

    # Create required paths if they don't exist
    dc_day_path.mkdir(parents=True, exist_ok=True)

    dc_delay_report_file = f"Delay Report {dc_selected_team_leader} {dc_day_dir}.xlsx"
    dc_delay_report_path = dc_day_path / dc_delay_report_file

    # Handle Template Copy & Rename
    if dc_delay_report_path.exists():
        status_msg = f"Delay Report {dc_selected_team_leader} {dc_day_dir} already exists!\n{dc_day_path}"
        pass
        return False, status_msg  # Indicate failure and return
    else:
        print(f"Creating Delay Report {dc_selected_team_leader}")
        copy_and_rename_template(
            paths["templates"] / DELAY_TEMPLATE,
            dc_day_path,
            dc_delay_report_file,
        )

        cp_wb = load_workbook(filename=construction_wp_path, read_only=True)
        cp_ws = cp_wb["Const. Plan"]
        dc_delay_wb = load_workbook(filename=dc_delay_report_path)
        dc_delay_ws = dc_delay_wb.active

        copy_from_cp_to_delay(cp_ws, dc_delay_ws, tl_num, dc_day_dir)
        fill_delay_ws_cells(dc_delay_ws, cp_ws, tl_num)

        dc_delay_wb.save(dc_delay_report_path)
        return True, None  # Indicate success and return


def copy_and_rename_template(src_path, dest_path, new_name):
    """
    Copy a source file to a destination and rename it.
    """

    shutil.copy(src_path, dest_path / src_path.name)
    (dest_path / src_path.name).rename(dest_path / new_name)


def fill_delay_ws_cells(delay_ws, cp_ws, team_leader_index):
    """
    Fill specific cells of a delay worksheet with pre-defined values or patterns.
    """

    username = username_entry.get()
    cells_to_fill = {
        (8, 8): username,
        (16, 5): "Foreman",
        (17, 5): "Team Leader",
        (16, 7): "SEMI",
        (17, 7): "SEMI",
        (28, 2): "Y",
        (29, 2): "Y",
        (
            8,
            6,
        ): f"{cp_ws.cell(row=int(team_leader_index), column=7).value} to {cp_ws.cell(row=int(team_leader_index),
                                                                                     column=8).value}",
        (
            8,
            4,
        ): f"{cp_ws.cell(row=int(team_leader_index), column=9).value} - "
           f"{cp_ws.cell(row=int(team_leader_index), column=10).value}",
    }
    for (row, col), value in cells_to_fill.items():
        set_cell(delay_ws, row, col, value)

    # Set fill patterns for specific cells
    pattern_fill_cells = ["B3", "G7", "C7", "B5", "B6", "F8", "B8", "F5", "F6", "H8"]
    for cell in pattern_fill_cells:
        delay_ws[cell].fill = PatternFill(bgColor="FFFFFF")


def set_cell(wb_sheet, row, column, value, fill=None):
    """
    Set a specific cell's value and optionally its fill pattern in a workbook sheet.

    Args:
        wb_sheet: Target workbook sheet.
        row (int): Row number of the cell.
        column (int): Column number of the cell.
        value: The value to be set for the cell.
        fill (optional): Fill pattern for the cell.
    """

    """Utility function to set cell values and, optionally, a fill pattern."""
    cell = wb_sheet.cell(row=row, column=column)
    cell.value = value
    if fill:
        cell.fill = PatternFill(bgColor=fill)


def copy_from_cp_to_delay(cp_ws, delay_ws, team_leader_num, day_folder):
    """
    Copy data from a construction plan worksheet to a delay worksheet.

    Args:
        cp_ws: Source construction plan worksheet.
        delay_ws: Target delay worksheet.
        team_leader_num (int): Index of the team leader to consider.
        day_folder (str): String representation of the day folder.
    """

    mapping = {
        # (delay_row, delay_col): (cp_col, transform_fn)
        (3, 2): (None, lambda _: day_folder),
        (5, 6): (None, lambda _: ""),
        (6, 6): (None, lambda _: ""),
        (7, 7): (13, None),
        (7, 3): (11, None),
        (5, 2): (5, None),
        (6, 2): (6, None),
        (32, 2): (22, None),
        (34, 2): (23, None),
        (33, 2): (24, None),
        (16, 1): (11, None),
        (17, 1): (13, None),
        (8, 2): (21, None),
    }

    for delay_coords, (cp_col, transform) in mapping.items():
        delay_row, delay_col = delay_coords
        if cp_col is None:
            value = transform(None)
        else:
            value = cp_ws.cell(row=int(team_leader_num), column=cp_col).value
            if transform:
                value = transform(value)
        set_cell(delay_ws, delay_row, delay_col, value)


def clear_cells():
    """
    Clear all the entry cells defined in the ENTRIES_CONFIG and reset related global variables.
    """

    global ENTRIES_CONFIG

    # Dynamically get the entries using their names
    entries = [globals()[entry_name] for entry_name in ENTRIES_CONFIG.keys()]

    # Clear all the entries
    for entry in entries:
        if not entry == frame4_reason_entry:
            entry.delete(0, "end")
        else:
            frame4_reason_entry.set("")

    # Reset variables
    frame4_workers_var.set(0)
    frame4_vehicles_var.set(0)


def get_cell_mapping():
    """
    Generate a mapping of cell configurations based on the global ENTRIES_CONFIG.

    Globals:
        Uses ENTRIES_CONFIG to generate the mapping.

    Returns:
        dict: Mapping of widget configurations with row, column, and optional time_format details.
    """

    mapping = {}
    for entry_name, config in ENTRIES_CONFIG.items():
        mapping[globals()[entry_name]] = {
            "row": config["row"],
            "col": config["col"],
            "time_format": config.get("time_format", False),
        }
    return mapping


def load_delay_wb():
    """
    Load the delay report workbook and populate certain GUI components based on its contents

    Note:
        If a workbook is already open, it closes it before loading a new one.
    """

    global delay_report_wb, delay_report_path, delay_report_ws

    def insert_value(row, col, widget, time_format=False):
        cell_value = delay_report_ws.cell(row=row, column=col).value
        if cell_value:
            if time_format and not isinstance(cell_value, str):
                cell_value = cell_value.strftime("%H:%M")
            if isinstance(cell_value, str):
                widget.insert(0, cell_value)

    try:
        delay_report_wb.close()

        delay_wb_name = delay_report_path.name.replace(".xlsx", "")
        print(f"Previous file closed: {delay_wb_name}")

        delay_report_path = delay_report_path.parent
    except AttributeError:
        pass

    delay_report_path = get_delay_report_path_for_tl(team_leader_name)
    delay_report_wb = load_workbook(filename=delay_report_path)
    delay_report_ws = delay_report_wb["Sheet1"]

    mapping = get_cell_mapping()
    for widget, details in mapping.items():
        row = details["row"]
        col = details["col"]
        time_format = details.get("time_format", False)
        insert_value(row, col, widget, time_format)


def populate_listbox():
    """
    Populate dm_tl_listbox with team leader names present in the delays_dir_path directory.
    Filenames are sorted by their stem (name without extension).
    """
    global tl_names_dict

    # Check if the listbox is populated
    if dm_tl_listbox.size() > 0:
        # Clear the listbox
        dm_tl_listbox.delete(0, END)

    try:
        tl_names_dict = {}

        for child in sorted(delays_dir_path.iterdir(), key=lambda x: x.stem):
            if child.is_file():
                tl_name = child.stem

                # Extract the team leader's name using regex
                match = re.search(
                    r"Delay Report ([\w\s]+(?: \+ [\w\s]+)?) \d{2}\.\d{2}\.\d{2}", tl_name
                )
                if match:
                    leader_name = match.group(1)
                    tl_names_dict[leader_name] = tl_name

        # Populate the listbox with only the keys (leader names)
        for leader_name in tl_names_dict.keys():
            dm_tl_listbox.insert(END, leader_name)

    except FileNotFoundError:
        pass


def construct_delay_report_path(tl_name=None):
    """
    Construct and return the path to the delay report based on the selected date
    and optionally, the team leader's name.
    """

    global delays_dir_path

    dm_selected_date = pd.Timestamp(dm_dates_combobox.get())
    formatted_dates, week = derive_dates(dm_selected_date)
    m_formatted_date = formatted_dates["dot"]
    paths = define_related_paths()
    dir_path = (
            paths["delays"] / str(dm_selected_date.year) / f"WW{week}" / m_formatted_date
    )
    if tl_name:
        return dir_path / f"Delay Report {tl_name} {dir_path.name}.xlsx"
    else:
        return dir_path


def dm_combo_selected(event):
    """
    Handle the event when a new date is selected in the dm_dates_combobox.
    Updates the global delays directory path and repopulates the listbox with
    relevant filenames. It also enables the required configuration buttons.
    """
    global delays_dir_path

    delays_dir_path = construct_delay_report_path()
    populate_listbox()

    # Set configurations
    set_config(save_button, state="normal")


def get_selected_item_from_listbox():
    cs = dm_tl_listbox.curselection()
    if not cs:
        return None
    name = dm_tl_listbox.get(cs[0])
    return name


def dm_on_tl_listbox_2_click(event):
    """
    Handle the event when a team leader name in the listbox is double-clicked.
    Sets the displayed team leader name, clears previous data, loads new data,
    and sets line status.
    """
    global team_leader_name

    tl_name_selected.config(text=get_selected_item_from_listbox())
    team_leader_name = get_selected_item_from_listbox()
    team_leader_name = tl_names_dict[team_leader_name]

    if not team_leader_name:
        return
    print(f"Loading : {team_leader_name}")
    clear_cells()
    load_delay_wb()
    line_status()


def get_delay_report_path_for_tl(team_leader):
    return delays_dir_path / f"{team_leader}.xlsx"


# noinspection PyTypeChecker
def dm_on_tl_listbox_rename(event):
    """
    Handle the event when a team leader name in the listbox is right double-clicked.
    Allows the user to rename a team leader and updates the related Excel file accordingly.
    """
    global team_leader_name

    try:
        team_leader_name = get_selected_item_from_listbox()
        team_leader_name = tl_names_dict[team_leader_name]

    except KeyError:
        messagebox.showwarning("Warning", "Please select the Team Leader and try again")
        return

    print(team_leader_name)

    print(f"Renaming {team_leader_name}")
    # Request the new team leader name
    new_team_leader_name = simpledialog.askstring("Input", "Enter the new Team leader name:", parent=app)
    if new_team_leader_name:
        new_team_leader_name = new_team_leader_name.strip()
    else:
        return

    # Confirmation of new file name
    new_delay_report_path = construct_delay_report_path(new_team_leader_name)
    confirm = messagebox.askyesno(
        "Confirmation",
        f"Old name : {team_leader_name}\n"
        f"New name : {new_delay_report_path.name[:-5]}\n\n"
        f"Are you sure you want to rename?"
    )
    if not confirm:
        return

    # Update the Excel cell with the new team leader's name
    delay_ws = delay_report_wb["Sheet1"]
    set_cell(delay_ws, 7, 7, new_team_leader_name)
    set_cell(delay_ws, 17, 1, new_team_leader_name)

    # Save changes and rename the file
    temp_delay_report_path = get_delay_report_path_for_tl(team_leader_name)
    delay_report_wb.save(temp_delay_report_path)

    if not new_delay_report_path.exists():
        temp_delay_report_path.rename(new_delay_report_path)
        messagebox.showinfo("Success", f"File renamed successfully!")
    else:
        messagebox.showwarning("Warning", "A file with that name already exists!")
        return

    team_leader_name = new_team_leader_name

    # Clear the listbox and repopulate it
    populate_listbox()

    # Convert tuple to list
    list_items = list(dm_tl_listbox.get(0, "end"))

    # Find the index of the new team leader
    new_tl_index = None
    for i, item in enumerate(list_items):
        if new_team_leader_name in item:
            new_tl_index = i
            break

    # If the new team leader's name is found, select that item
    if new_tl_index is not None:
        dm_tl_listbox.selection_set(new_tl_index)
        dm_on_tl_listbox_2_click(None)  # Passing None or a dummy event
    else:
        print(f"{new_team_leader_name} not found in the list box.")


def dm_on_tl_listbox_delete(event):
    """
    Handle the event to delete a selected item from the listbox and the actual file.
    """

    # Get the selected item from the listbox
    cs = dm_tl_listbox.curselection()
    if not cs:
        return
    selected_item = dm_tl_listbox.get(cs[0])
    selected_item = tl_names_dict[selected_item]

    # Construct the full path to the file
    file_path = delays_dir_path / f"{selected_item}.xlsx"

    # Confirm deletion with the user
    confirm = messagebox.askyesno(
        "Confirmation", f"Do you really want to delete {selected_item}?"
    )
    if not confirm:
        return

    # Delete the file
    if file_path.exists():
        file_path.unlink()
        # Remove the item from the listbox
        dm_tl_listbox.delete(cs)
        print(f"Deleted: {file_path.name}")
    else:
        messagebox.showwarning("Warning", f"{selected_item} not found!")


def save_delay_wb():
    """
    Save the details from the GUI entries into the delay workbook related to the
    selected team leader.
    """
    global delay_report_path

    if not team_leader_name:
        return

    temp_delay_report_wb = load_workbook(filename=delay_report_path)
    temp_delay_report_ws = temp_delay_report_wb["Sheet1"]

    # Check for empty worker entries and update w1_entry if necessary
    if all([globals()[entry_name].get() == "" for entry_name in WORKER_ENTRIES]):
        globals()["w1_entry"].delete(0, "end")  # Clear existing content first
        globals()["w1_entry"].insert(0, ".")

    # Check for empty vehicle entry and update if necessary
    if globals()["v1_entry"].get() == "":
        globals()["v1_entry"].delete(0, "end")  # Clear existing content first
        globals()["v1_entry"].insert(0, "No vehicle")

    # Direct assignments using ENTRIES_CONFIG
    for entry_name, config in ENTRIES_CONFIG.items():
        cell_address = config["cell"]
        entry = globals()[entry_name]
        temp_delay_report_ws[cell_address] = entry.get()

    temp_delay_report_wb.save(delay_report_path)
    # clear_cells()
    # load_delay_wb()
    line_status()
    print(f"Saved successfully : {team_leader_name}")


def status_check():
    """
    Checks if all necessary conditions are met and updates the status displayed in the frame.
    Status will be set to "Completed" if all criteria are met, otherwise "Not completed".
    """
    global status_color

    if (
            start_time == 1
            and end_time == 1
            and reason_var == 1
            and worker1_var == 1
            and vehicle1_var == 1
    ):
        set_config(frame3_status, text="Completed", bootstyle="success")

        status_color = 1
    else:
        set_config(frame3_status, text="Not completed", bootstyle="danger")
        status_color = 0


def set_entry_status(entry, var_name, default_val=0):
    """
    Updates the style of a given entry based on its content and modifies a global variable accordingly.
    """
    if entry.get() == "" and not entry == frame4_reason_entry:
        entry.config(style="danger.TEntry")
        globals()[var_name] = default_val
        frame4_reason_entry.config(style="danger.TCombobox")
        globals()["reason_var"] = default_val

    else:
        if not entry == frame4_reason_entry and not entry == frame4_reason_entry:
            entry.config(style="success.TEntry")
            globals()[var_name] = 1
            frame4_reason_entry.config(style="success.TCombobox")
            globals()["reason_var"] = 1


def line_status():
    """
    Iterates over the pre-defined entry widgets to set their styles and states based on their contents.
    Also triggers the overall status check for the frame.
    """
    for entry_name, config in ENTRIES_CONFIG.items():
        entry = globals()[entry_name]
        var_name = config["var"]
        set_entry_status(entry, var_name)

    if globals()["w1_entry"].get() == "" and frame4_workers_var.get() == 0:
        for entry_name in WORKER_ENTRIES:
            entry = globals()[entry_name]
            entry.config(style="danger.TEntry")
        globals()["worker1_var"] = 0
    else:
        for entry_name in WORKER_ENTRIES:
            entry = globals()[entry_name]
            entry.config(style="success.TEntry")
        globals()["worker1_var"] = 1

    status_check()


def set_config(widget, **options):
    """
    Configures properties for a given tkinter widget.
    """
    widget.config(**options)


def extract_date_from_path(path):
    """
    Extracts date and week information from a given directory path.
    """
    str_date = path.name  # The last component of the path (should be the date)

    week_info = (
        path.parent.name
    )  # The second last component (should be 'Working Week Nxx')

    # Extract week number from week_info
    week_num = week_info.split("N")[-1]  # Assumes format 'Working Week Nxx'

    # Convert the string date to a datetime object
    dt_date = datetime.strptime(str_date, "%d.%m.%y")

    return str_date, dt_date, week_num


def extract_src_path_from_date(str_date, dt_date, week_num):
    """
    Constructs source file paths based on provided date information.

    Returns:
    - tuple: A tuple containing daily report path and weekly delay path.
    """
    paths, c_formatted_dates, p_formatted_dates = derive_paths_from_date(dt_date)

    # Creating the WW Delay Table
    weekly_delay_name = f"Weekly Delay table {week_num}.xlsx"
    weekly_delay_f_path = delays_dir_path.parent
    weekly_delay_path = weekly_delay_f_path / weekly_delay_name

    # Creating the CIIM Daily Report Table file path
    daily_report_name = derive_report_name(str_date)
    daily_report_f_path = paths["day"]
    daily_report_path = daily_report_f_path / daily_report_name

    print(daily_report_path)
    print(weekly_delay_path)

    return daily_report_path, weekly_delay_path


def are_files_locked(src_filepath: Path, dest_filepath: Path) -> bool:
    return is_file_locked(src_filepath) or is_file_locked(dest_filepath)


def is_file_locked(filepath: Path) -> bool:
    locked = None
    file_object = None
    if filepath.exists():
        try:
            # Try to open and close the file in append mode.
            # If this fails, the file is locked.
            file_object = filepath.open("a")
            if file_object:
                locked = False
        except IOError:
            locked = True
        finally:
            if file_object:
                file_object.close()
    return locked


def format_time(time_obj):
    if time_obj:
        return time_obj.strftime("%H:%M")
    else:
        return "None"


def transfer_data_to_cancelled(source_file, destination_file, mappings):
    """
    Transfers data from a source file to a destination file based on column mappings provided.
    """
    # Load the workbooks and worksheets in read_only mode for the source file

    src_wb = load_workbook(source_file, read_only=True)
    src_ws = src_wb["Const. Plan"]

    dest_wb = load_workbook(destination_file)
    dest_ws = dest_wb["Work Cancelled"]

    # Print all headers from the source file
    print("Source headers:", [cell.value for cell in src_ws[2]])
    print("Destination headers:", [cell.value for cell in dest_ws[3]])

    src_header = {
        cell.value: col_num + 1
        for col_num, cell in enumerate(src_ws[2])
        if cell.value in mappings
           or any(cell.value in key for key in mappings if isinstance(key, tuple))
    }

    dest_header = {
        cell.value: col_num + 1
        for col_num, cell in enumerate(dest_ws[3])
        if cell.value in mappings.values()
    }

    dest_row_counter = 4
    observation_col = src_header.get("Observations", None)

    transferred_rows = 0
    for row_num, row in enumerate(src_ws.iter_rows(min_row=3, values_only=True), 4):
        if observation_col:

            # Check if the row is blank by looking at certain key columns
            key_column_indexes = [
                src_header['T.P Start [Time]'] - 1,
                src_header['Team Leader\nName (Phone)'] - 1,
                src_header['Date [DD/MM/YY]'] - 1,
            ]
            if all(not row[idx] for idx in key_column_indexes):
                continue  # Skip the row as it is considered blank

            observation_value = row[observation_col - 1]  # -1 because row is 0-indexed
            # It will skip the rows that are blank or those who doesn't have 'Cancel' in the Observation cell
            if not observation_value or "cancel" not in observation_value.lower():
                continue
            # It will skip the rows that were cancelled by OCS/Scada/TS
            if any(word in observation_value.lower() for word in ["scada", "ocs", "ts"]):
                print(f"Skipping row {row_num} due to observation value: {observation_value}")
                continue

        for src_col, dest_col in mappings.items():
            if isinstance(src_col, tuple) and src_col == (
                    "T.P Start [Time]",
                    "T.P End [Time]",
            ):
                start_time_col = src_header.get("T.P Start [Time]")
                end_time_col = src_header.get("T.P End [Time]")

                # Debugging checks:
                if not start_time_col:
                    print(f"'T.P Start [Time]' not found in source header for row {row_num}.")
                if not end_time_col:
                    print(f"'T.P End [Time]' not found in source header for row {row_num}.")
                if dest_col not in dest_header:
                    print(f"'{dest_col}' not found in destination header for row {row_num}.")

                if start_time_col and end_time_col and dest_col in dest_header:
                    ww_start_time = format_time(row[start_time_col - 1])
                    ww_end_time = format_time(row[end_time_col - 1])

                    # should be checked to determine if the row is effectively "empty"
                    if row[observation_col - 1]:
                        combined_time = f"{ww_start_time}-{ww_end_time}"
                        dest_ws.cell(
                            row=dest_row_counter, column=dest_header[dest_col]
                        ).value = combined_time
                        print(f"Writing combined time to row {dest_row_counter} in destination.")
                    else:
                        # Break out of the loop if it finds an empty key cell
                        break
                else:
                    print(f"Missing columns in source or destination for row {row_num}.")

            else:
                if src_col in src_header and dest_col in dest_header:
                    dest_ws.cell(
                        row=dest_row_counter, column=dest_header[dest_col]
                    ).value = row[src_header[src_col] - 1]
                else:
                    print(
                        f"Missing columns '{src_col}' or '{dest_col}' in source or destination for row {row_num}.")

        dest_row_counter += 1
        transferred_rows += 1

    dest_wb.save(destination_file)
    dest_wb.close()
    src_wb.close()

    return transferred_rows


def transfer_cancelled_wrapper():
    # --------------------- Logic Handling Functions ---------------------

    def on_cancel():
        top_level.destroy()

    def on_confirm():

        try:
            cancelled_transferred = transfer_data_to_cancelled(
                construction_wp_path,
                weekly_delay_path,
                TO_CANCELLED_MAPPING)

            # Updated message to show how many rows were transferred
            transferred_message = f"{cancelled_transferred} rows transferred." if cancelled_transferred is not None \
                else "No rows were transferred."
            messagebox.showinfo("Success", transferred_message)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

        top_level.destroy()

    # ------------------------- Main Function -------------------------

    messagebox.showwarning("Reminder", f"Make sure to fill first all the Cancelled works in the Construction plan!")

    initial_dir = CIIM_DIR_PATH / 'General Updates' / 'Delays+Cancelled works' / str(datetime.now().year)
    weekly_delay_path = filedialog.askopenfilename(initialdir=initial_dir)
    weekly_delay_path = Path(weekly_delay_path)

    if not weekly_delay_path:
        return

    # Root
    top_level = ttk.Toplevel()
    top_level.withdraw()  # Hide the window initially
    top_level.title("Transfer Cancelled works")
    top_level.geometry('320x200')
    top_level.resizable(False, False)

    # Center the top_level window
    center_window(top_level, top_level.master)
    top_level.deiconify()  # Show the window after centering

    confirm_frame = ttk.Frame(top_level)
    confirm_frame.pack(fill="both", expand=True)

    confirm_frame.grid_rowconfigure(0, weight=1)
    confirm_frame.grid_rowconfigure(7, weight=1)
    confirm_frame.grid_columnconfigure(4, weight=1)

    source_label = ttk.Label(confirm_frame, text=f"SOURCE: {construction_wp_path.name}")
    source_label.grid(row=1, column=0, columnspan=5, pady=10, padx=20, sticky="nsew")

    destination_label = ttk.Label(confirm_frame, text=f"DESTINATION: {weekly_delay_path.name}")
    destination_label.grid(row=2, column=0, columnspan=5, padx=20, pady=5, sticky="w")

    # Toolbar + Buttons
    toolbar_confirm_frame = ttk.Frame(master=confirm_frame)
    toolbar_confirm_frame.grid(row=7, columnspan=5, sticky="nsew")

    confirm_transfer_button = ttk.Button(toolbar_confirm_frame, text="Confirm", command=on_confirm, width=10)
    confirm_transfer_button.pack(side=RIGHT, anchor="se", padx=5, pady=10)

    cancel_button = ttk.Button(toolbar_confirm_frame, text="Cancel", command=on_cancel, width=10, style="secondary")
    cancel_button.pack(side=RIGHT, anchor="se", padx=5, pady=10)


def transfer_delay_data(source_file, destination_file, mappings, dest_start_row=4):
    """
    Transfers data from a source file to a destination file based on column mappings provided.
    Skips rows where 'Observations' column contains 'cancel'.
    """

    # Load the workbooks and worksheets
    src_wb = load_workbook(source_file, read_only=True)
    src_ws = src_wb.active

    dest_wb = load_workbook(destination_file)
    dest_ws = dest_wb["Work Delay"]

    # Mapping source and destination headers to their respective column numbers
    src_header = {
        cell.value: col_num + 1
        for col_num, cell in enumerate(src_ws[3])
        if cell.value in mappings or any(cell.value in key for key in mappings if isinstance(key, tuple))
    }

    dest_header = {
        cell.value: col_num + 1
        for col_num, cell in enumerate(dest_ws[3])
        if cell.value in mappings.values()
    }

    # Find the column number for "Observations" in the source file
    observation_col_num = None
    for col_num, cell in enumerate(src_ws[3]):
        if cell.value == "Observations":
            observation_col_num = col_num + 1
            break

    if observation_col_num is None:
        print("Warning: 'Observations' column not found in the source file.")
        return

    # Print headers for debugging
    print("Source headers:", [cell.value for cell in src_ws[3]])
    print("Destination headers:", [cell.value for cell in dest_ws[3]])

    dest_row_counter = dest_start_row
    transferred_rows = 0

    # Iterating through each row in the source worksheet
    for row_num, row in enumerate(src_ws.iter_rows(min_row=4, values_only=True), 4):
        # Checks if the row is blank by looking at certain key columns
        key_column_indexes = [
            src_header['T.P Start [Time]'] - 1,
            src_header['Team Leader\nName (Phone)'] - 1,
            src_header['Date [DD/MM/YY]'] - 1
        ]
        if all(not row[idx] for idx in key_column_indexes):
            continue  # Skip the row as it is considered blank

        if observation_col_num and row[observation_col_num - 1] and "cancel" in row[observation_col_num - 1].lower():
            print(f"Skipping row {row_num} due to 'cancel' in Observations value : {row[observation_col_num - 1]}")
            continue

        # Transferring data based on mappings
        for src_col, dest_col in mappings.items():
            if src_col in src_header and dest_col in dest_header:
                dest_ws.cell(row=dest_row_counter, column=dest_header[dest_col]).value = row[
                    src_header[src_col] - 1]

            else:
                print(f"Missing columns '{src_col}' or '{dest_col}' in source or destination for row {row_num}.")

        dest_row_counter += 1
        transferred_rows += 1  # Increment only if data is actually transferred in this row

    # Save and close workbooks
    dest_wb.save(destination_file)
    dest_wb.close()
    src_wb.close()

    return transferred_rows


def transfer_delay_wrapper():
    # ----------------------- Logic Handling Functions -----------------------

    def on_confirm():

        delay_transferred = 0

        delay_value = delay_entry.get().strip()
        delay_int = int(delay_value) if delay_value else None

        try:
            # Transfer for delay
            if delay_int is not None:
                delay_transferred = transfer_delay_data(
                    daily_report_path,
                    weekly_delay_path,
                    TO_DELAY_MAPPINGS,
                    delay_int)

            # Updated message to show how many rows were transferred
            transferred_message = f"{delay_transferred} rows transferred." if delay_transferred is not None else ("No "
                                                                                                                  "rows were transferred.")
            messagebox.showinfo("Success", transferred_message)
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

        top_level.destroy()

    def on_cancel():
        top_level.destroy()

    def on_back():
        confirm_frame.pack_forget()
        input_frame.pack(fill="both", expand=True)

    def validate_input(char):
        # Validation function to allow only numeric input
        return char.isdigit() or char == ""

    def on_next():

        # Check if files are locked
        if are_files_locked(daily_report_path, weekly_delay_path):
            messagebox.showwarning(
                "File Locked",
                f"Please close the following Excel files before proceeding:\n{daily_report_path.name}\n"
                f"{weekly_delay_path.name}",
            )
            top_level.destroy()
            return

        input_frame.pack_forget()  # Hide input frame
        confirm_frame.pack(fill="both", expand=True)  # Show confirm frame

        delay_input = delay_entry.get().strip()

        delay_input_label.config(text=f"Delay row: {delay_input}")

    # Function to update the state of the "Confirm" button based on the Entry widgets' content
    def update_next_button_state(*args):
        delay_input = delay_entry.get().strip()

        if delay_input:
            next_button["state"] = "normal"
        else:
            next_button["state"] = "disabled"

    # ----------------------- Main Function -----------------------

    # Makes sure a date is selected in the Combo box first
    try:
        str_date, dt_date, week_num = extract_date_from_path(delays_dir_path)
        daily_report_path, weekly_delay_path = extract_src_path_from_date(str_date, dt_date, week_num)

    except ValueError:
        messagebox.showerror("Error", "You need to choose a date first!")
        return

    messagebox.showinfo("Reminder", f"Make sure that {daily_report_path.name}\nis filled before proceeding.")

    # Root
    top_level = ttk.Toplevel()
    top_level.withdraw()  # Hide the window initially
    top_level.title("Transfer to Weekly")
    top_level.geometry('320x200')
    top_level.resizable(False, False)

    # Center the top_level window
    center_window(top_level, top_level.master)
    top_level.deiconify()  # Show the window after centering

    # ------ Input Frame------
    input_frame = ttk.Frame(master=top_level)
    input_frame.pack(fill="both", expand=True)

    input_frame.grid_rowconfigure(0, weight=1)
    input_frame.grid_rowconfigure(3, weight=1)
    input_frame.grid_columnconfigure(2, weight=1)

    explain_label = ttk.Label(input_frame, text="Enter the row number to which you want to transfer")
    explain_label.grid(row=0, columnspan=3, padx=20, pady=10, sticky="nsew")

    delay_label = ttk.Label(input_frame, text="Delay row")
    delay_label.grid(row=1, column=0, padx=20, pady=5, sticky="w")

    delay_entry = ttk.Entry(input_frame, width=12)
    delay_entry.grid(row=1, padx=5, column=1, pady=5, sticky="w")

    # Toolbar + Buttons
    toolbar_input_frame = ttk.Frame(master=input_frame)
    toolbar_input_frame.grid(row=3, columnspan=3, sticky="nsew")

    next_button = ttk.Button(toolbar_input_frame, text="Next >", command=on_next, width=10, state="disabled")
    next_button.pack(side=RIGHT, anchor="se", padx=5, pady=10)

    cancel_button = ttk.Button(toolbar_input_frame, text="Cancel", command=on_cancel, width=10, style="secondary")
    cancel_button.pack(side=RIGHT, anchor="se", padx=5, pady=10)

    # ------ Confirm Frame------
    confirm_frame = ttk.Frame(top_level)

    confirm_frame.grid_rowconfigure(0, weight=1)
    confirm_frame.grid_rowconfigure(7, weight=1)
    confirm_frame.grid_columnconfigure(4, weight=1)

    source_label = ttk.Label(confirm_frame, text=f"SOURCE: {daily_report_path.name}")
    source_label.grid(row=1, column=0, columnspan=5, pady=10, padx=20, sticky="nsew")

    destination_label = ttk.Label(confirm_frame, text=f"DESTINATION: {weekly_delay_path.name}")
    destination_label.grid(row=2, column=0, columnspan=5, padx=20, pady=5, sticky="w")

    delay_input_label = ttk.Label(confirm_frame, text="Delay row: ")
    delay_input_label.grid(row=3, column=0, padx=20, pady=5, sticky="w")

    # Bind the Entry widgets to the update function to be called whenever their content changes
    delay_entry.bind("<KeyRelease>", update_next_button_state)

    # Apply the validation function to the Entry widgets
    vcmd = top_level.register(validate_input)
    delay_entry.config(validate="key", validatecommand=(vcmd, '%S'))

    # Toolbar + Buttons
    toolbar_confirm_frame = ttk.Frame(master=confirm_frame)
    toolbar_confirm_frame.grid(row=7, columnspan=5, sticky="nsew")

    confirm_transfer_button = ttk.Button(toolbar_confirm_frame, text="Confirm", command=on_confirm, width=10)
    confirm_transfer_button.pack(side=RIGHT, anchor="se", padx=5, pady=10)

    back_button = ttk.Button(toolbar_confirm_frame, text="< Back", command=on_back, width=10, style="secondary")
    back_button.pack(side=RIGHT, anchor="se", padx=5, pady=10)


def derive_dates(selected_date):
    """
    Derive all related paths from a given date including multiple date formats.
    """

    day, month, year = [
        selected_date.strftime(pattern) for pattern in ["%d", "%m", "%Y"]
    ]
    week = selected_date.strftime(
        "%U"
    )  # returns the week number considering the first day of the week as Sunday

    formatted_dates = {
        "slash": f"{day}/{month}/{year[-2:]}",
        "dot": f"{day}.{month}.{year[-2:]}",
        "compact": f"{year[-2:]}{month}{day}",
    }

    return formatted_dates, week


def derive_paths_from_date(selected_date):
    """
    Constructs various related paths based on a given date.
    """

    c_day, c_month, c_year = [
        selected_date.strftime(pattern) for pattern in ["%d", "%m", "%Y"]
    ]
    p_date_datetime = selected_date - timedelta(days=1)
    p_day, p_month, p_year = [
        p_date_datetime.strftime(pattern) for pattern in ["%d", "%m", "%Y"]
    ]

    c_week = selected_date.strftime(
        "%U"
    )  # returns the week number considering the first day of the week as Sunday

    c_formatted_dates = {
        "slash": f"{c_day}/{c_month}/{c_year[-2:]}",
        "dot": f"{c_day}.{c_month}.{c_year[-2:]}",
        "compact": f"{c_year[-2:]}{c_month}{c_day}",
    }

    p_formatted_dates = {
        "slash": f"{p_day}/{p_month}/{p_year[-2:]}",
        "dot": f"{p_day}.{p_month}.{p_year[-2:]}",
        "compact": f"{p_year[-2:]}{p_month}{p_day}",
    }

    paths = {
        "year": CIIM_DIR_PATH / f"Working Week {c_year}",
        "week": CIIM_DIR_PATH / f"Working Week {c_year}" / f"Working Week N{c_week}",
        "day": CIIM_DIR_PATH
               / f"Working Week {c_year}"
               / f"Working Week N{c_week}"
               / f"{c_year[-2:]}{c_month}{c_day}",
        "previous_year": CIIM_DIR_PATH / f"Working Week {p_year}",
        "previous_week": CIIM_DIR_PATH
                         / f"Working Week {p_year}"
                         / f"Working Week N{c_week}",
        "previous_day": CIIM_DIR_PATH
                        / f"Working Week {p_year}"
                        / f"Working Week N{c_week}"
                        / f"{p_year[-2:]}{p_month}{p_day}",
    }

    return paths, c_formatted_dates, p_formatted_dates


def pick_date():
    """
    Prompt user to select a date using the Query box widget.
    - Updates the GUI (button's text) to display the chosen week and date.
    - Checks if a directory corresponding to the chosen date already exists.
    - Depending on directory existence, it updates the state of entry widgets and the create button.
    """

    global fc_selected_date
    cal = Querybox()
    fc_selected_date = cal.get_date(bootstyle="danger")
    paths, c_formatted_dates, p_formatted_dates = derive_paths_from_date(fc_selected_date)

    day_message_exist = f'{c_formatted_dates["compact"]} folder already exists'
    if paths["day"].exists():
        messagebox.showerror("Error", day_message_exist)
        return

    # Feedback using button's text
    calendar_button.config(
        text=f"WW: {fc_selected_date.strftime('%U')}     Date: {fc_selected_date.strftime('%d.%m.%Y')} "
    )

    entries_state = "disabled" if paths["day"].exists() else "normal"
    set_config(ocs_entry, state=entries_state)
    set_config(scada_entry, state=entries_state)
    set_config(create_button, state=entries_state)
    calendar_button.config(bootstyle="success.Outline")

    return paths


def check_and_create_path(path):
    """
    Checks and creates a directory for the given path if it doesn't exist.
    """

    if not path.exists():
        path.mkdir(parents=True, exist_ok=True)


def derive_report_name(date, template="CIIM Report Table {}.xlsx"):
    """
    Derive a report filename based on the given date.

    Args:
        date (str): The date used for naming.
        template (str, optional): String template for report naming. Default is "CIIM Report Table {}.xlsx".

    Returns:
        str: Report name with the date inserted into the template.
    """

    return template.format(date)


def create_folders_for_entries(path, entry, prefix):
    """
    Create a set of folders based on the provided entry value.
    Each folder will have a unique name prefixed by the given prefix and will contain subfolders named "Pictures" and
    "Worklogs".
    """

    """Utility to create folders for the given prefix and entry."""
    for i in range(int(entry.get() or 0)):
        (path / f"{prefix}{i + 1}" / "Pictures").mkdir(parents=True, exist_ok=True)
        (path / f"{prefix}{i + 1}" / "Worklogs").mkdir(parents=True, exist_ok=True)


def create_folders():
    """
    Execute the process to:
    - Import paths and formatted dates.
    - Create main paths for year, week, and day.
    - Notify the user when a folder is successfully created.
    - Generate, copy, and rename report files.
    - Create additional folders based on entry values.
    - Create other necessary folders.
    - Reset and configure GUI widgets.
    - Handle data report writing and copying.
    """

    # Importing the paths and the formatted dates
    paths, c_formatted_dates, p_formatted_dates = derive_paths_from_date(
        fc_selected_date
    )
    main_paths = define_related_paths()

    # Creating main paths
    for key in ["year", "week", "day"]:
        Path(paths[key]).mkdir(parents=True, exist_ok=True)

    if paths["day"].exists():
        day_created_message = (
            f'{c_formatted_dates["compact"]} folder was created successfully!'
        )
        messagebox.showinfo(None, day_created_message)

    # Derive report name and handle file copying and renaming
    ciim_daily_report = derive_report_name(c_formatted_dates["dot"])
    print(f"Generated report name: {ciim_daily_report}")

    templates_path = main_paths["templates"]
    fc_ciim_template_path = templates_path / DAILY_REPORT_TEMPLATE

    # Copy and rename
    print(f'Copying template to: {paths["day"]}')
    shutil.copy(fc_ciim_template_path, paths["day"])

    new_report_path = paths["day"] / ciim_daily_report
    print(f"Renaming file to: {new_report_path}")
    template_in_dest = paths["day"] / DAILY_REPORT_TEMPLATE
    if template_in_dest.exists():
        template_in_dest.rename(new_report_path)

        # Introduce a slight delay
        time.sleep(1)
    else:
        print(f'Template not found in {paths["day"]}!')

    # Creating folders for entries
    create_folders_for_entries(paths["day"], ocs_entry, "W")
    create_folders_for_entries(paths["day"], scada_entry, "S")

    # Creating other necessary folders
    folders_to_create = [
        "Foreman",
        "Track possession",
        "TS Worklogs",
        "PDF Files",
        "Worklogs",
    ]
    for folder in folders_to_create:
        (paths["day"] / folder).mkdir(exist_ok=True)

    # Handle data report writing and copying
    if Path(construction_wp_path).parent != Path(paths["week"]):
        print("Not copying works to the selected date")
        return

    write_data_to_report(
        construction_wp_path,
        c_formatted_dates["slash"],
        paths["day"],
        TO_DAILY_REPORT_MAPPINGS,
    )

    # Only show the popup if previous day path exists
    if paths["previous_day"].exists():
        result = messagebox.askyesno(
            title=None,
            message=f"Transfer data to {derive_report_name(p_formatted_dates['dot'])}?",
        )
        print(result)
        if result is True:
            write_data_to_previous_report(
                construction_wp_path,
                p_formatted_dates["slash"],
                paths["previous_day"],
                TO_DAILY_REPORT_MAPPINGS,
                paths["previous_day"] / derive_report_name(p_formatted_dates["dot"]),
            )
            return

    # Reset and configure other widgets
    ocs_entry.delete(0, END)
    scada_entry.delete(0, END)
    set_config(ocs_entry, state="disabled")
    set_config(scada_entry, state="disabled")
    set_config(create_button, state="disabled")
    calendar_button.config(text="Browse", bootstyle="danger.Outline")


def write_data_to_excel(src_path, target_date, target_directory, mappings, start_row=4):
    """
    Write data from the source Excel to a target report based on given mappings.
    """

    target_datetime = pd.to_datetime(target_date, format="%d/%m/%y", errors="coerce")
    formatted_target_date = target_datetime.strftime("%d.%m.%y")
    report_filename = derive_report_name(formatted_target_date)
    target_report_path = Path(target_directory / report_filename)

    usecols_value = [mappings[header] for header in mappings.keys()]
    df = pd.read_excel(src_path, skiprows=1, usecols=usecols_value)
    print(df.columns)

    df["Date [DD/MM/YY]"] = pd.to_datetime(
        df["Date [DD/MM/YY]"], format="%d/%m/%Y", dayfirst=True, errors="coerce"
    )
    target_df = df[df["Date [DD/MM/YY]"] == target_datetime]

    # Open the target workbook
    target_workbook = load_workbook(filename=target_report_path)
    target_worksheet = target_workbook.active

    try:

        # Write headers (using the mappings keys as headers)
        for col, header in enumerate(mappings.keys(), 2):  # Starting from column B
            target_worksheet.cell(row=start_row - 1, column=col, value=header)

        # Write data
        for row_idx, (index, row_data) in enumerate(target_df.iterrows(), start=start_row):
            for col_idx, header in enumerate(mappings.keys(), 2):  # Starting from column B
                target_worksheet.cell(row=row_idx, column=col_idx, value=row_data[header])

        # Change the value of cell O3 to "Work Description"
        target_worksheet['O3'] = "Work Description"

        target_workbook.save(target_report_path)
        print(f"Report for {formatted_target_date} has been updated and saved.")

    except ValueError as e:
        messagebox.showerror("Error", f"Failed to read Excel file: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def write_data_to_report(src_path, target_date, target_directory, mappings):
    """
    Write data to the current day's report.
    """

    write_data_to_excel(src_path, target_date, target_directory, mappings)


def write_data_to_previous_report(
        src_path, target_date, target_directory, mappings, target_report_path
):
    """
    Write data to the previous day's report. Prompt the user to select a starting row.
    """
    # Prompt the user for the starting row

    while is_file_locked(target_report_path):
        response = messagebox.askretrycancel(
            "File Locked", f"Please close {target_report_path.name} and try again!"
        )
        if not response:  # If user selects "Cancel"
            return
    # Continue with the rest of the function after this block

    start_row_delay = simpledialog.askinteger(
        "Input", "Enter the starting row:", minvalue=4
    )
    start_row_delay = int(start_row_delay)
    # If the user cancels the prompt or doesn't enter a valid number, exit the function
    if not start_row_delay:
        return

    write_data_to_excel(
        src_path, target_date, target_directory, mappings, start_row=start_row_delay
    )


def is_dark_theme(theme):
    # Define what constitutes a dark theme
    dark_themes = ['darkly', 'superhero', 'cyborg', 'solar', 'vapor']
    return theme in dark_themes


def change_theme():
    """Changes the theme of the app to the selected theme."""
    current_theme = theme_var.get()
    style.theme_use(current_theme)  # Set the theme
    style.configure("TButton", font=("Roboto", 9, "bold"))
    style.configure("TMenubutton", font=("Roboto", 9, "bold"))
    style.configure("success.Link.TButton", font=("Roboto", 9, "bold"), anchor=W)
    print(current_theme)

    update_icons(current_theme)


def show_frame(frame_name):
    global current_frame

    for name, frame in frames.items():
        frame.pack_forget()
        if name == frame_name:
            frame.pack(fill="both", expand=True)
            current_frame = frame_name


def load_icon(icon_path):
    # Load the icon image
    img = Image.open(icon_path)

    # Convert the image to a format that Tkinter can use
    tk_img = ImageTk.PhotoImage(img)
    return tk_img


def update_icons(theme):
    # Update the icons based on the theme
    icon_set = 'light' if is_dark_theme(theme) else 'dark'

    # Keep a reference to the icons in a global scope
    global photo_images
    photo_images = {}

    # Iterate over the notebook's tabs
    for tab_name, frame in {'Home': tab1, 'File': tab2, 'Folder': tab3, 'Edit': tab4}.items():
        icon_path = notebook_tabs[tab_name][icon_set]
        photo_images[tab_name] = load_icon(icon_path)  # Assuming load_icon returns an ImageTk.PhotoImage
        my_notebook.tab(frame, image=photo_images[tab_name])


def clock():
    time = datetime.now()
    hour = time.strftime(" %H:%M ")
    weekday = time.strftime("%A")
    day = time.day
    month = time.strftime("%m")
    year = time.strftime("%Y")

    hour_label.configure(text=hour)
    hour_label.after(6000, clock)

    day_label.configure(text=weekday + ", " + str(day) + "/" + str(month) + "/" + str(year))


def show_context_menu(event):
    """
    Show the context menu on right-click.
    """
    cs = dm_tl_listbox.curselection()
    if cs:
        context_menu.post(event.x_root, event.y_root)


def edit_username():
    # Get the current username
    current_username = username_var.get()

    # Ask the user for a new username using a simple dialog
    new_username = simpledialog.askstring(parent=tab1, title="Edit Username", prompt="Enter new username:",
                                          initialvalue=current_username)

    # If the user provides a new username (i.e., didn't cancel the dialog), update the username_var
    if new_username is not None:
        username_var.set(new_username)


def enable_transfer_button(event):
    global transfer_button_visible

    if not transfer_button_visible:
        transfer_button.pack(side=RIGHT, padx=5, pady=10)
        transfer_button_visible = True
    else:
        transfer_button.pack_forget()
        transfer_button_visible = False


def show_notebook(frame_name):
    if construction_wp_var.get() != "" and username_var.get() != "":
        show_frame(frame_name)
    else:
        messagebox.showerror(title="Error", message="Please fill your name and the construction work plan")


def update_edit_frame_based_on_tab_change(event):
    """
    Resets the cells and the entries if tab is changed
    """
    selected_tab_index = event.widget.index("current")

    # Checking whenever the tab is not "Edit"
    if not selected_tab_index == 4:
        # Check if the listbox has any selected item
        tl_name_selected.config(text="None")
        clear_cells()
        line_status()
        populate_listbox()

        for entry_name in ENTRIES_CONFIG.keys():
            entry = globals()[entry_name]
            if not entry == frame4_reason_entry:
                entry.config(style="default.TEntry")
            else:
                frame4_reason_entry.config(style="default.TCombobox")


def display_dist_list():
    show_frame("Dist list")

    paths = define_related_paths()
    proc_path = paths["procedure"]

    if proc_path.exists():
        try:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(proc_path, sheet_name='Dist. List', usecols='B, D, F, H')
            df.fillna('', inplace=True)

            # Iterate over the DataFrame and the text widgets at the same time
            for col, text_widget in zip(df.columns[:4], text_widgets):
                # Clear the text widget first
                text_widget.delete('1.0', END)
                # Insert the data into the text widget
                column_data = '\n'.join(df[col].astype(str))
                text_widget.insert('1.0', column_data)
                # Highlight lines containing "cc" after inserting the text

                highlight_lines_containing_cc(text_widget)


        except ValueError as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")


# Function to toggle between template and original content
def toggle_content(text_widget, template, original_contents, column):
    current_content = text_widget.get('1.0', 'end-1c')
    if current_content.strip() == template.strip():
        # If current content is the template, replace with original email
        text_widget.delete('1.0', 'end')
        text_widget.insert('1.0', original_contents[column])
    else:
        # If current content is not the template, store it and insert template
        original_contents[column] = current_content
        text_widget.delete('1.0', 'end')
        text_widget.insert('1.0', template)

    # Reapply the highlight to the text widget
    highlight_lines_containing_cc(text_widget)


def highlight_lines_containing_cc(text_widget):
    # Defines a tag for 'email' and 'whatsapp'
    text_widget.tag_configure("highlight", underline=True, font=("Roboto", 9, "bold"))
    # Defines a tag for 'cc' with a different style
    text_widget.tag_configure("cc_highlight", font=("Roboto", 9, "bold"), background='#f90b31')

    words_to_highlight = ["email", "whatsapp", "preview"]

    # Iterate over the list of words and highlight them with the 'highlight' tag
    for word in words_to_highlight:
        start_index = '1.0'
        while True:
            start_index = text_widget.search(word, start_index, 'end', nocase=True)
            if not start_index:
                break
            end_index = f"{start_index} lineend"
            text_widget.tag_add("highlight", start_index, end_index)
            start_index = f"{end_index}+1c"

    # Search and highlight 'cc' with a different tag 'cc_highlight'
    start_index = '1.0'
    while True:
        start_index = text_widget.search("cc", start_index, 'end', nocase=True)
        if not start_index:
            break
        end_index = f"{start_index} lineend"
        text_widget.tag_add("cc_highlight", start_index, end_index)
        start_index = f"{end_index}+1c"


def display_phone_list():
    show_frame("Phones")

    # Ideally, you should also handle potential errors here, such as the file not existing.
    phones_df = pd.read_csv('names.csv')

    team_leader_phones = phones_df["Team Leader Name"]
    foreman_phones = phones_df['Foreman Name']

    # Convert the series to a single string with line breaks, excluding NaN values.
    tl_phones_str = "\n".join(team_leader_phones.dropna().astype(str))
    foreman_phones_str = "\n".join(foreman_phones.dropna().astype(str))

    # Update the contents of the Text widgets.
    tl_phones_list.delete("1.0", "end")
    tl_phones_list.insert("end", tl_phones_str)

    foreman_list.delete("1.0", "end")
    foreman_list.insert("end", foreman_phones_str)


def copy_to_clipboard(event, text_widget):
    try:
        # Get the current line index
        current_index = text_widget.index(CURRENT)
        line = current_index.split('.')[0]
        # Extract the line's content
        line_text = text_widget.get(f"{line}.0", f"{line}.end")
        # Clear the clipboard and append new content
        phones_frame.clipboard_clear()
        phones_frame.clipboard_append(line_text.strip())
        # Optionally, show a message that the content was copied
        messagebox.showinfo("Info", f"Copied to clipboard: {line_text.strip()}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def open_precedure_file():
    paths = define_related_paths()
    proc_path = paths["procedure"]
    os.startfile(proc_path)


def open_wp_file():
    global construction_wp_path
    os.startfile(construction_wp_path)


def open_faults():
    paths = define_related_paths()
    faults_path = paths["faults"]
    os.startfile(faults_path)


def open_passdown():
    paths = define_related_paths()
    passdown_path = paths["passdown"]

    files = sorted(passdown_path.glob("*.xlsx"), key=os.path.getmtime, reverse=True)
    filename = files[0]

    print(f"Latest passdown : {filename.stem}")

    os.startfile(filename)


# ========================= Root config =========================
pyglet.font.add_file('digital-7/digital-7.ttf')

app = ttk.Window()
windll.shcore.SetProcessDpiAwareness(1)
app.resizable(0, 0)
app.title("Smart CIIM")

# Grid
app.grid_columnconfigure(0, weight=1)
app.grid_rowconfigure(0, weight=1)
# Geometry
app_width = 750
app_height = 550
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
x = (screen_width / 2) - (app_width / 2)
y = (screen_height / 2) - app_height
app.geometry(f"{app_width}x{app_height}+{int(x)}+{int(y)}")

# app.iconbitmap(bitmap='images/snake.ico')
# app.iconbitmap(default='images/snake.ico')
# ============================ Style ============================
style = ttk.Style()
style.configure("TButton", font=("Roboto", 9, "bold"), )
style.configure("success.Link.TButton", font=("Roboto", 9, "bold"), anchor=W)
style.configure("TMenubutton", font=("Roboto", 9, "bold"))

# =========================== Variables ===========================
username_var = StringVar()
username = ""
current_frame = None
# A global variable to keep track of whether the text widgets have been created
# Paths
CIIM_DIR_PATH = Path("/")
delays_dir_path = Path("/")
construction_wp_path = Path("/")
construction_wp_var = StringVar()
delay_report_path = Path("/")
selected_date = ""
# Tkinter variables
team_leader_name = ""
status_color = IntVar()
previous_day_entry = IntVar()
dc_day, dc_month, dc_week, dc_year = "", "", "", ""
start_time, end_time, reason_var, worker1_var, vehicle1_var = 0, 0, 0, 0, 0
dc_selected_date = ""
fc_selected_date = ""
# Lists and associated data
tl_list = []
cp_dates = []
tl_index = []
# Miscellaneous variables
dc_selected_team_leader = ""
tl_num = 0
delay_report_wb = ""
frame4_workers_var = IntVar()
frame4_vehicles_var = IntVar()
DELAY_TEMPLATE = "Delay Report template v.02.xlsx"
DAILY_REPORT_TEMPLATE = "CIIM Report Table v.1.xlsx"
# Themes
tl_names_dict = {}
theme_var = StringVar()
# Those TLs won't appear in the Listbox that creates delays
TL_BLACKLIST = [
    "Eliyau Ben Zgida",
    "Emerson Gimenes Freitas",
    "Emilio Levy",
    "Samuel Lakko",
    "Ofer Akian",
    "Wissam Hagay",
    "Rami Arami",
]
# List of Delay reasons
delay_reasons = [
    "Delay due to no TP",
    "Delay due to track vehicle maneuvers",
    "Delay due to waiting for the ISR/WSP Supervisor",
    "Delay due to waiting for the ISR Safety/ISR Comm. Supervisor",
    "Delay due to in the release of the electrified area/rail",
    "Delay due to coordination with the control center for the TP",
    "Delay due to track vehicle maneuvers",
    "Delay due to real hours are different for the 612",
    "--Other--"
]
# ============================ Mappings ============================
CONSTRUCTION_WP_HEADERS = [
    "Discipline [OCS/Old Bridges/TS/Scada]",
    "WW [Nº]",
    "Date [DD/MM/YY]",
    "T.P Start [Time]",
    "T.P End [Time]",
    "T.P Start [K.P]",
    "T.P End [K.P]",
    "EP",
    "ISR Start Section [Name]",
    "ISR  End Section [Name]",
    "Foremen [Israel]",
    "Team Name",
    "Team Leader\nName (Phone)",
    "Work Description (Baseline)",
    "ISR Safety Request",
    "ISR Comm&Rail:",
    "ISR T.P request (All/Track number)",
    "Observations"
]
HEADER_TO_INDEX = {header: index for index, header in enumerate(CONSTRUCTION_WP_HEADERS)}
# All the Headers from the Construction Work Plan match the CIIM Report Table
TO_DAILY_REPORT_MAPPINGS = {
    header: header for header in CONSTRUCTION_WP_HEADERS}

TO_DELAY_MAPPINGS = {
    "WW [Nº]": "WW",
    "Discipline [OCS/Old Bridges/TS/Scada]": "Discipline [OCS, Scada, TS]",
    "Date [DD/MM/YY]": "Date",
    "Delay details (comments + description)": "Reason",
    "Team Name": "Team Name",
    "Team Leader\nName (Phone)": "Team leader",
    "EP": "ISR section {EP}",
    "T.P Start [Time]": "TP Start Time (TAK)",
    "Work Description": "Work Description",
    "Actual Start Time (TL):": "Actual Start Time (Real Start time - TL)",
    "T.P End [Time]": "TP Finish Time (TAK)",
    "Actual Finish Time (TL):": "Actual Finish Time (Real Finish time - TL)",
    "Number of workers": "Workers",
}

TO_CANCELLED_MAPPING = {
    "WW [Nº]": "WW",
    "Discipline [OCS/Old Bridges/TS/Scada]": "Discipline [OCS, Scada, TS]",
    "Date [DD/MM/YY]": "Date",
    "Observations": "Reason",
    "Team Leader\nName (Phone)": "Team leader",
    "Work Description (Baseline)": "Work Description",
    ("T.P Start [Time]", "T.P End [Time]"): "Planned hour per shift",
    "EP": "ISR section {EP}",
}

# Centralized list of entries and their configurations
ENTRIES_CONFIG = {
    "frame4_stime_entry": {
        "cell": "F5",
        "var": "start_time",
        "time_format": True,
        "row": 5,
        "col": 6,
    },
    "frame4_endtime_entry": {
        "cell": "F6",
        "var": "end_time",
        "time_format": True,
        "row": 6,
        "col": 6,
    },
    "frame4_reason_entry": {"cell": "A11", "var": "reason_var", "row": 11, "col": 1},
    "w1_entry": {"cell": "A18", "var": "worker1_var", "row": 18, "col": 1},
    "w2_entry": {"cell": "A19", "var": "worker2_var", "row": 19, "col": 1},
    "w3_entry": {"cell": "A20", "var": "worker3_var", "row": 20, "col": 1},
    "w4_entry": {"cell": "A21", "var": "worker4_var", "row": 21, "col": 1},
    "w5_entry": {"cell": "A22", "var": "worker5_var", "row": 22, "col": 1},
    "w6_entry": {"cell": "A23", "var": "worker6_var", "row": 23, "col": 1},
    "w7_entry": {"cell": "A24", "var": "worker7_var", "row": 24, "col": 1},
    "w8_entry": {"cell": "A25", "var": "worker8_var", "row": 25, "col": 1},
    "v1_entry": {"cell": "D28", "var": "vehicle1_var", "row": 28, "col": 4},
}

WORKER_ENTRIES = [
    "w1_entry",
    "w2_entry",
    "w3_entry",
    "w4_entry",
    "w5_entry",
    "w6_entry",
    "w7_entry",
    "w8_entry",
]

# =========================== Frames ===========================
frames = {
    "Login": ttk.Frame(master=app),
    "Notebook": ttk.Frame(master=app),
    "Phones": ttk.Frame(master=app),
    "Dist list": ttk.Frame(master=app)
}

# ====================== Login Frame ======================
# ================== Background Image ===================
login_frame = frames["Login"]
bg = ImageTk.PhotoImage(file='images/background.png')
login_button_img = ImageTk.PhotoImage(file='images/button_img.png')

# Show image
label1 = Label(master=login_frame, image=bg)
label1.place(x=0, y=0)

username_entry = ttk.Entry(master=login_frame, textvariable=username_var, width=50, font=("Roboto", 11, "bold"),
                           style="light"
                           )
username_entry.place(x=210, y=170)

path_entry = ttk.Entry(master=login_frame, textvariable=construction_wp_var,
                       width=50, font=("Roboto", 11, "bold"), style="light")
path_entry.place(x=210, y=288)
path_entry.bind('<Button-1>', open_const_wp)
username_entry.bind('<Tab>', open_const_wp)

login_button = ttk.Button(
    master=login_frame,
    command=lambda: show_notebook("Notebook"),
    image=login_button_img, style="light"

)
login_button.place(x=250, y=380)

# ====================== Notebook Config ======================
my_notebook = ttk.Notebook(master=frames["Notebook"])
my_notebook.pack(fill="both", expand=True)

tab1 = ttk.Frame(master=my_notebook)
tab2 = ttk.Frame(master=my_notebook)
tab3 = ttk.Frame(master=my_notebook)
tab4 = ttk.Frame(master=my_notebook)

notebook_tabs = {
    'Home': {
        'light': 'images/icons8-home-24(6).png',
        'dark': 'images/icons8-home-24(5).png'
    },
    'File': {
        'light': 'images/icons8-file-24(3).png',
        'dark': 'images/icons8-file-24(2).png'
    },

    'Folder': {
        'light': 'images/icons8-folder-24(2).png',
        'dark': 'images/icons8-folder-24(1).png'
    },
    'Edit': {
        'light': 'images/icons8-edit-24(2).png',
        'dark': 'images/icons8-edit-24(1).png'
    }
}

# Dictionary to store the photo images to prevent garbage collection
photo_images = {}

my_notebook.add(child=tab1)
my_notebook.add(child=tab2)
my_notebook.add(child=tab3)
my_notebook.add(child=tab4)

# ====================== Tab 1 - Home ======================

tab1.columnconfigure(0, weight=1)
tab1.columnconfigure(1, weight=0)
tab1.rowconfigure(1, weight=1)
tab1.rowconfigure(2, weight=1)

time_frame = ttk.Frame(master=tab1)
time_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)

# Then, packing the user-related labels at the bottom
user_frame = ttk.Frame(master=tab1)
user_frame.grid(row=0, column=0, pady=5, sticky="nsew")
active_user_label = ttk.Label(user_frame, text="Active :")
active_user_label.pack(side=LEFT)

display_username = ttk.Label(user_frame, textvariable=username_var, bootstyle="info", font=("Roboto", 9, "bold"))
display_username.pack(side=LEFT)
display_username.bind("<Double-1>", lambda e: edit_username())
ToolTip(display_username, text='Double-click to rename')

# Packing the hour and day labels at the top first
hour_label = ttk.Label(master=time_frame, text="12:49", font="digital-7 90")
hour_label.pack()  # North/top alignment

day_label = ttk.Label(master=time_frame, text="Saturday 22/01/2023", font=("verdana", 20), style="secondary")
day_label.pack(padx=5, pady=5)  # North/top alignment

path_frame = ttk.Frame(master=tab1)
path_frame.grid(row=2, column=0, sticky='nsew', padx=5, pady=5)

home_browse_button = ttk.Button(master=path_frame, text='Change file', command=select_const_wp, bootstyle='secondary',
                                width=10)
home_browse_button.pack(anchor='sw', side='left', pady=5)

path_entry = ttk.Entry(master=path_frame, textvariable=construction_wp_var)
path_entry.pack(anchor='s', side='left', fill='x', expand=True, pady=5)

utilities_frame = ttk.Frame(master=tab1)
utilities_frame.grid(row=0, column=1, rowspan=3, sticky="nsew", padx=5, pady=5)  # Adjust grid placement as needed

# Utilities
open_wp_button = ttk.Button(master=utilities_frame, text="Open Construction Plan", command=lambda: open_wp_file(),
                            style="success.Link.TButton")
open_wp_button.pack(fill='x', padx=5, pady=5)
open_proc_button = ttk.Button(master=utilities_frame, text="Open Procedure", command=lambda: open_precedure_file(),
                              style="success.Link.TButton")
open_proc_button.pack(fill='x', padx=5, pady=5)

open_faults_button = ttk.Button(master=utilities_frame, text="Open Ele. Control Center ", command=lambda: open_faults(),
                                style="success.Link.TButton")
open_faults_button.pack(fill='x', padx=5, pady=5)
open_passdown_button = ttk.Button(master=utilities_frame, text="Open Passdown ",
                                  command=lambda: open_passdown(),
                                  style="success.Link.TButton")
open_passdown_button.pack(fill='x', padx=5, pady=5)

phones_button = ttk.Button(master=utilities_frame, text="Phone numbers", command=lambda: display_phone_list(),
                           bootstyle='link.success')
phones_button.pack(fill='x', padx=5, pady=5)

dist_button = ttk.Button(master=utilities_frame, text="Distribution List", command=lambda: display_dist_list(),
                         style='success.Link.TButton')
dist_button.pack(fill='x', padx=5, pady=5)

transfer_all_button = ttk.Button(master=utilities_frame, text="Transfer Cancelled ",
                                 command=lambda: transfer_cancelled_wrapper(),
                                 style="success.Link.TButton")
transfer_all_button.pack(fill='x', padx=5, pady=5)

# Create Theme menu option
theme_button = ttk.Menubutton(utilities_frame, text="Theme")
theme_button.pack(fill='x', padx=5, pady=5, side='bottom')

theme_menu = ttk.Menu(theme_button)

for theme_name in style.theme_names():
    theme_menu.add_radiobutton(label=theme_name, variable=theme_var, command=change_theme)
# Associates the inside menu with the menubutton
theme_button["menu"] = theme_menu

# ====================== Tab 1 -Phones frame ======================

phones_frame = frames["Phones"]
phones_frame.rowconfigure(2, weight=1)
phones_frame.columnconfigure(0, weight=1)
phones_frame.columnconfigure(2, weight=1)

# Team Leader names on the left side
tl_label = ttk.Label(master=phones_frame, text="Team Leaders", anchor="center", font=("Roboto", 9, "bold"))
tl_label.grid(row=1, column=0, columnspan=2, pady=5, sticky="nsew")

tl_phones_list = ttk.Text(master=phones_frame, wrap="word", spacing1=7)
tl_phones_list.grid(row=2, column=0, sticky="nsew")
tl_phones_scroll = ttk.Scrollbar(master=phones_frame, style="round", command=tl_phones_list.yview)
tl_phones_scroll.grid(row=2, column=1, sticky="nsw")
tl_phones_list.config(yscrollcommand=tl_phones_scroll.set)

# Foreman names on the right side
foreman_label = ttk.Label(master=phones_frame, text="Foremen", anchor="center", font=("Roboto", 9, "bold"))
foreman_label.grid(row=1, column=2, columnspan=2, padx=5, pady=5, sticky="nsew")
foreman_list = ttk.Text(master=phones_frame, wrap="word", spacing1=7)
foreman_list.grid(row=2, column=2, sticky="nsew")
foreman_scroll = ttk.Scrollbar(master=phones_frame, style="round", command=foreman_list.yview)
foreman_scroll.grid(row=2, column=3, sticky="nsw")
foreman_list.config(yscrollcommand=foreman_scroll.set)
tl_phones_list.config(cursor="hand2")
foreman_list.config(cursor="hand2")

ToolTip(tl_phones_list, "Click to Copy", delay=500)
ToolTip(foreman_list, "Click to Copy", delay=500)

# Back button
phone_back_button = ttk.Button(master=phones_frame, text="< Back", command=lambda: show_frame("Notebook"), width=10,
                               bootstyle="secondary")

phone_back_button.grid(row=0, columnspan=6, padx=10, pady=10, sticky="w")

# Bindings
tl_phones_list.bind("<Button-1>", lambda event: copy_to_clipboard(event, tl_phones_list))
foreman_list.bind("<Button-1>", lambda event: copy_to_clipboard(event, foreman_list))

# ====================== Tab 1 - Dist. list frame ======================

dist_frame = frames["Dist list"]

# Configure the frame to give equal weight to all columns
for i in range(4):
    dist_frame.columnconfigure(i, weight=1)
dist_frame.rowconfigure(2, weight=1)

templates = {
    "Pass down": "Hi Dana,\n\nNothing special happened during the shift.",
    "Preview": "              Email (SEMI):"
               "\n\nHi all,"
               "\n\n  1. TLs ... didn't send forms."
               "\n  2.TLs ... didn't send worklogs."
               "\n  3.TLs ... was delayed due to no TP."
               "\n\n\n\n\n\n\n             Email (ISR):"
               "\n\nHi Yoni,"
               "\n\nFind attached the draft of the CIIM Report.",
    "Not Approved": "              Email (12:00):"
                    "\n\nHi Randall,"
                    "\n\nFind attached the updated plan for tonight (dd.mm.yy) and tomorrow morning (dd.mm.yy) / the "
                    "weekend (dd-dd.mm.yy)."
                    "\nPlease add the WSP supervisors, ISR working charges and ISR communication supervisors names in "
                    "the file."
                    "\n\n\n           Whatsapp (16:00):"
                    "\n\nGood afternoon everyone,\nAttached is the updated work file for tonight (dd.mm.yy) and "
                    "tomorrow morning (dd.mm.yy)."
                    "\nPlease note that the hours listed are the starting hours of the T.P. Please keep in touch with "
                    "your managers about the time you should be in the field."
                    "\nGood luck."
                    "\n*TPs and supervisors in charge will be updated by ISR as soon as possible.*",
    "Approved": "      Email (17:00~20:00):"
                "\n\nHi all,"
                "\n\nPlease find the approved Construction Plan for tonight (dd.mm.yy) and tomorrow morning ("
                "dd.mm.yy) / the weekend  (dd-dd.mm.yy)."
}

# Store the original content of the text widgets
original_contents = ['' for _ in range(4)]

# Text widgets list
text_widgets = [Text(dist_frame) for _ in range(4)]


# Function to create button commands
def make_command(col, tw, temp):
    return lambda: toggle_content(tw, temp, original_contents, col)


# Creates buttons and text widgets, and place them in the frame inside the canvas
for column, (label_text, template) in enumerate(templates.items()):
    button = ttk.Button(dist_frame, text=label_text, command=make_command(column, text_widgets[column], template),
                        bootstyle="link")
    button.grid(row=1, column=column, pady=5, padx=2)
    text_widget = text_widgets[column]
    text_widget.grid(row=2, column=column, sticky="nsew", padx=2)
    ToolTip(button, text="Click for template", delay=600)

# Back button
dist_back_button = ttk.Button(master=dist_frame, text="< Back", command=lambda: show_frame("Notebook"), width=10,
                              bootstyle="secondary")
dist_back_button.grid(row=0, columnspan=4, padx=10, pady=10, sticky="w")

# ====================== Tab 2 - File ======================

tab2.rowconfigure(0, weight=1)
tab2.columnconfigure(0, weight=1)
tab2.columnconfigure(2, weight=1)

tab2_mid_frame = ttk.Frame(master=tab2)
tab2_mid_frame.grid(row=0, column=1, sticky="nsew")

tab2_mid_frame.rowconfigure(1, weight=1)

dc_select_date_label = ttk.Label(master=tab2_mid_frame, text="   Select date:  ", )
dc_select_date_label.grid(row=1, column=1, padx=5, pady=5, sticky="e")
dates_combobox = ttk.Combobox(master=tab2_mid_frame, values=cp_dates, postcommand=update_combo_list)
dates_combobox.set("Date")
dates_combobox.bind("<<ComboboxSelected>>", dc_combo_selected)
dates_combobox.grid(row=1, column=2, padx=5, pady=5, sticky="w")
dc_tl_listbox = Listbox(master=tab2_mid_frame, border=5, selectmode=ttk.EXTENDED, height=20, width=40)
dc_tl_listbox.bind("<Return>", dc_on_listbox_create)
dc_tl_listbox.grid(row=2, column=1, columnspan=2, pady=20)
tab2_scrollbar = ttk.Scrollbar(master=tab2_mid_frame, style="round", command=dc_tl_listbox.yview)
tab2_scrollbar.grid(row=2, column=3, pady=20, sticky="nsw")
dc_tl_listbox.config(yscrollcommand=tab2_scrollbar.set)
dc_create_button = ttk.Button(master=tab2_mid_frame, text="Create", command=dc_on_listbox_create, width=10,
                              state=DISABLED)
dc_create_button.grid(row=3, column=1, pady=10, sticky="e")

dc_create_all_button = ttk.Button(master=tab2_mid_frame, text="Create all", command=create_all_delays, width=10,
                                  style="outline",
                                  state=DISABLED)
dc_create_all_button.grid(row=3, column=2, pady=10, sticky="e")

# ====================== Tab 3 - Folder ======================
tab3.rowconfigure(0, weight=1)
tab3.columnconfigure(0, weight=1)
tab3.columnconfigure(2, weight=1)

tab3_mid_frame = ttk.Frame(master=tab3)
tab3_mid_frame.grid(row=0, column=1, sticky='nsew')

select_folder_label = ttk.Label(master=tab3_mid_frame, text="   Select date:  ")
select_folder_label.grid(row=0, column=0, padx=5, pady=34, sticky="e")
calendar_button = ttk.Button(master=tab3_mid_frame, text="Browse", command=pick_date, width=23, style="danger.Outline")
calendar_button.grid(row=0, column=1, padx=5, pady=5, sticky="w")

discipline_frame = ttk.Frame(master=tab3_mid_frame)
discipline_frame.grid(row=1, column=0, sticky="nsew", columnspan=2, pady=40)
fc_ocs_label = ttk.Label(master=discipline_frame, text="Num of OCS works")
fc_ocs_label.grid(row=0, column=0, sticky="e", padx=5, pady=30, )
fc_scada_label = ttk.Label(master=discipline_frame, text="Num of SCADA works")
fc_scada_label.grid(row=1, column=0, sticky="e", padx=5, pady=5)
ocs_entry = ttk.Entry(master=discipline_frame, state="disabled", width=10)
ocs_entry.grid(row=0, column=1, sticky="e", padx=5)
scada_entry = ttk.Entry(master=discipline_frame, state="disabled", width=10)
scada_entry.grid(row=1, column=1, sticky="e", padx=5)

# Button
create_button = ttk.Button(master=tab3_mid_frame, text="Create", command=create_folders, state="disabled", width=10)
create_button.grid(row=2, column=0, columnspan=2, sticky="n", pady=10)

# ====================== Tab 4 - Edit ======================
tab4.rowconfigure(0, weight=1)
tab4.rowconfigure(1, weight=1)
tab4.columnconfigure(1, weight=1)

# Date Select
menu3_frame1 = ttk.Frame(master=tab4)
menu3_frame1.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
dc_select_date_label = ttk.Label(menu3_frame1, text="   Select date:  ")
dc_select_date_label.pack(side="left")
dm_dates_combobox = ttk.Combobox(menu3_frame1, values=cp_dates, postcommand=update_combo_list)
dm_dates_combobox.set("Date")
dm_dates_combobox.bind("<<ComboboxSelected>>", dm_combo_selected)
dm_dates_combobox.pack(side="left")

# Team Leaders Listbox
menu3_frame2 = ttk.Frame(master=tab4)
menu3_frame2.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
dm_tl_listbox = Listbox(menu3_frame2, border=5)

dm_tl_listbox.pack(side=LEFT, fill=BOTH, expand=True)
dm_tl_listbox.bind("<Double-1>", dm_on_tl_listbox_2_click)

tab3_scrollbar = ttk.Scrollbar(master=menu3_frame2, style="round", command=dm_tl_listbox.yview)
tab3_scrollbar.pack(side=RIGHT, fill=BOTH)
dm_tl_listbox.config(yscrollcommand=tab3_scrollbar.set)

# Create a context menu
context_menu = Menu(master=app, tearoff=0)
context_menu.add_command(label="Rename", command=lambda: dm_on_tl_listbox_rename(None))
context_menu.add_command(label="Delete", command=lambda: dm_on_tl_listbox_delete(None))

dm_tl_listbox.bind("<Button-3>", show_context_menu)
dm_tl_listbox.bind("<Delete>", dm_on_tl_listbox_delete)

# It will clear the cells and update the listbox when changing to another tab
my_notebook.bind("<<NotebookTabChanged>>", update_edit_frame_based_on_tab_change)
# Frame 3 - Name + Status
menu3_frame3 = ttk.LabelFrame(master=tab4, text="Information", labelanchor="n", style="info")
menu3_frame3.grid(row=0, column=1, sticky="nsew", padx=10, pady=5)
ttk.Label(menu3_frame3, text="   Selected:").grid(row=0, column=0, sticky="w", pady=5)
tl_name_selected = ttk.Label(menu3_frame3, text="None", width=41, font=("Roboto", 9, "bold"))
tl_name_selected.grid(row=0, column=1, sticky="e")
tl_name_selected.bind("<Double-1>", dm_on_tl_listbox_rename)
ToolTip(tl_name_selected, text="Double-click to rename")

ttk.Label(menu3_frame3, text="Status:").grid(row=0, column=2, sticky="we", pady=5)
frame3_status = ttk.Label(menu3_frame3, text="Not completed", style="danger", font=("Roboto", 9, "bold"))
frame3_status.grid(row=0, column=3, sticky="e")

# Frame 4 - Manager
menu3_frame4 = ttk.Frame(master=tab4)

menu3_frame4.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
menu3_frame4.columnconfigure(0, weight=2)
menu3_frame4.columnconfigure(1, weight=1)
menu3_frame4.columnconfigure(2, weight=2)
menu3_frame4.columnconfigure(3, weight=1)
menu3_frame4.columnconfigure(4, weight=1)
menu3_frame4.rowconfigure(12, weight=1)

ttk.Label(menu3_frame4, text="Start time").grid(row=0, column=0, padx=5)
frame4_stime_entry = ttk.Entry(menu3_frame4)
frame4_stime_entry.grid(row=0, column=1, pady=2, sticky="w")

ttk.Label(menu3_frame4, text="End time").grid(row=1, column=0, padx=5)
frame4_endtime_entry = ttk.Entry(menu3_frame4)
frame4_endtime_entry.grid(row=1, column=1, pady=2, sticky="w")

ttk.Label(menu3_frame4, text="Reason").grid(row=2, column=0, padx=5, pady=2)
frame4_reason_entry = ttk.Combobox(menu3_frame4, values=delay_reasons, takefocus=False)
frame4_reason_entry.grid(row=2, column=1, sticky="we", pady=2, columnspan=3)

sep = ttk.Separator(master=menu3_frame4)
sep.grid(row=3, column=1, columnspan=3, sticky="we", pady=5, )

# Workers
ttk.Label(menu3_frame4, text="Workers").grid(row=4, column=0, padx=5)
for i, entry_name in enumerate(WORKER_ENTRIES, start=4):
    globals()[entry_name] = ttk.Entry(menu3_frame4)
    globals()[entry_name].grid(row=i, column=1, sticky="w", pady=2)

# Vehicles
ttk.Label(menu3_frame4, text="Vehicles").grid(row=4, column=2, padx=5, sticky="e")
v1_entry = ttk.Entry(menu3_frame4)
v1_entry.grid(row=4, column=3, sticky="e")

# Toolbar Frame
toolbar_frame = ttk.Frame(master=tab4)
toolbar_frame.grid(row=2, column=0, columnspan=2, sticky="nsew")

# Button
save_button = ttk.Button(toolbar_frame, text="Save", command=save_delay_wb, state="disabled", width=10)
save_button.pack(anchor="n", side=RIGHT, padx=10, pady=10)

transfer_button_visible = False
transfer_button = ttk.Button(toolbar_frame, text="Transfer", command=transfer_delay_wrapper,
                             width=10, style="info")

# Bind the "End" key press event to enable_transfer_button
app.bind("<KeyPress-End>", enable_transfer_button)

show_frame("Notebook")
update_icons("dafault")
clock()

app.mainloop()

# TODO : Make that the user is updated when delay is managed and not created
