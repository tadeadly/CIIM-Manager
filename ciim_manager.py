import os
import sys
import re
import shutil
import datetime as dt
from datetime import datetime
from pathlib import Path
from tkinter import *
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from PIL import ImageTk, Image
import ttkbootstrap as ttk
from ttkbootstrap.tooltip import ToolTip
from ttkbootstrap.utility import enable_high_dpi_awareness



def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))

    return os.path.join(base_path, relative_path)


def define_related_paths():
    """Defines all paths relative to the global CIIM_FOLDER_PATH."""

    paths = {
        "Construction": base_path / "CIIM - General",
        "Faults": base_path / "CIIM - Faults" / "Fault Report Database.xlsx",
        "Templates": base_path / "CIIM - Guidelines" / "Templates",
        "Guidelines": base_path / "CIIM - Guidelines",
        "Procedure": base_path / "CIIM - Guidelines" / "Protocols" / "CIIM Procedure.xlsx",
        "Tracking": base_path / "CIIM - Admin Records" / "CIIM" / "Performance Tracking 2025",
        "Admin":  base_path / "CIIM - Admin Records"
    }

    return paths


def get_base_path_from_file(file_path):
    """Retrieve the CIIM folder path from the given file path."""
    return file_path.parent.parent.parent.parent


def select_const_wp():
    """
    Opens a file dialog for the user to select an Excel file.
    """
    global wp_var, wp_path, path_entry
    pattern = "WW*Construction Work Plan*.xlsx"
    path = filedialog.askopenfilename(filetypes=[("Excel Files", pattern)])

    if path:  # Check if a path was actually selected
        wp_path = Path(path)
        wp_var.set(wp_path.name)  # Set the StringVar to just the filename
        # After path has been set, update the dates
        update_dates_based_on_file()
        # If no path was selected, simply do nothing

        path_entry.configure(bootstyle="secondary")
        print(wp_path)

    return wp_path if path else None


def update_dates_based_on_file():
    global wp_path, base_path, cp_dates, ww_var

    if not wp_path or wp_path == Path("/"):
        return

    wb = load_workbook(filename=wp_path)
    base_path = get_base_path_from_file(wp_path)
    cp_dates = extract_unique_dates_from_worksheet(wb["Const. Plan"])

    if not cp_dates:
        messagebox.showwarning("No Dates Found", "No dates found in the construction plan.")
        wb.close()
        return

    str_date = cp_dates[0]
    formatted_str_date, dt_date, week_num = extract_date(str_date)

    # Remove any leading zeros and convert to integer
    week_num = int(str(week_num).lstrip("0") or "0")
    print(f"Week Number Extracted: {week_num}")

    ww_var.set(week_num)
    update_menu_labels()
    wb.close()



def extract_unique_dates_from_worksheet(sheet_name):
    """
    Extract unique dates from a given worksheet column
    """
    unique_dates = set()
    for cell in sheet_name["C"]:
        date_value_str = process_date_cell(cell)
        if date_value_str:
            unique_dates.add(date_value_str)

    return sorted(list(unique_dates))


def process_date_cell(cell):
    """
    Processes a given cell's value to extract the date
    """
    if not cell.value:
        return None

    if isinstance(cell.value, datetime):
        return cell.value.date().isoformat()

    try:
        date_value = datetime.strptime(cell.value, "%d/%m/%Y").date()
        return date_value.isoformat()
    except ValueError:
        return None


def extract_date(date_str):

    dt_date = datetime.strptime(date_str, "%Y-%m-%d")
    formatted_str_date = dt_date.strftime("%Y-%m-%d")
    week_num = calculate_week_num(dt_date)

    return formatted_str_date, dt_date, week_num


def extract_src_path_from_date(dt_date):
    """
    Constructs source file paths based on provided date information.
    """
    paths, c_formatted_dates,_  = derive_paths_from_date(dt_date)

    # Creating the CIIM Daily Report Table file path
    daily_report_name = derive_report_name(c_formatted_dates["dot"])
    daily_report_f_path = paths["day"]
    daily_report_path = daily_report_f_path / daily_report_name
    print(daily_report_path)

    return daily_report_path


def are_files_locked(src_filepath: Path, dest_filepath: Path) -> bool:
    return is_file_locked(src_filepath) or is_file_locked(dest_filepath)


def is_file_locked(filepath: Path) -> bool:
    locked = None
    file_object = None
    if filepath.exists():
        try:
            # Try to open and close the file in append mode.
            # If this fails, the file is locked.
            file_object = filepath.open("a")
            if file_object:
                locked = False
        except IOError:
            locked = True
        finally:
            if file_object:
                file_object.close()
    return locked


def transfer_data_to_cancelled(source_file, destination_file, mappings, dest_start_row):
    """
    Transfers data from a source file to a destination file based on column mappings provided.
    """
    src_wb = load_workbook(source_file, read_only=True)
    src_ws = src_wb["Const. Plan"]

    dest_wb = load_workbook(destination_file)
    dest_ws = dest_wb["Cancellations"]

    # Print all headers from the source file
    # print("Source headers:", [cell.value for cell in src_ws[2]])
    # print("Destination headers:", [cell.value for cell in dest_ws[2]])

    src_header = {
        cell.value: col_num + 1
        for col_num, cell in enumerate(src_ws[2])
        if cell.value in mappings
           or any(cell.value in key for key in mappings if isinstance(key, tuple))
    }

    dest_header = {
        cell.value: col_num + 1
        for col_num, cell in enumerate(dest_ws[2])
        if cell.value in mappings.values()
    }

    dest_row_counter = dest_start_row
    observation_col = src_header.get("Observations", None)

    transferred_rows = 0
    for row_num, row in enumerate(src_ws.iter_rows(values_only=True), 2):
        if observation_col:

            # Check if the row is blank by looking at certain key columns
            key_column_indexes = [
                src_header['T.P Start [Time]'] - 1,
                src_header['Team Leader\nName (Phone)'] - 1,
                src_header['Date'] - 1,
            ]
            # List of keywords to check in the observation column
            keywords = ["cancel", "activity moved", "completed", "done", "postponed"]

            if all(not row[idx] for idx in key_column_indexes):
                continue  # Skip the row as it is considered blank

            observation_value = row[observation_col - 1]  # -1 because row is 0-indexed

            # Skip rows that are blank or do not contain any of the specified keywords
            if not observation_value or not any(keyword in observation_value.lower() for keyword in keywords):
                continue

            # It will skip the rows that were cancelled by OCS/Scada/TS
            # if any(word in observation_value.lower() for word in ["by scada", "by ocs", "by ocs-l", "by ocs-d"]):
            #     print(f"Skipping row {row_num} due to observation value: {observation_value}")
            #     continue

        for src_col, dest_col in mappings.items():

            if src_col in src_header and dest_col in dest_header:
                dest_ws.cell(
                    row=dest_row_counter, column=dest_header[dest_col]
                ).value = row[src_header[src_col] - 1]
            else:
                print(
                    f"Missing columns '{src_col}' or '{dest_col}' in source or destination for row {row_num}.")

        dest_row_counter += 1
        transferred_rows += 1

    dest_wb.save(destination_file)
    src_wb.close()

    return transferred_rows


def transfer_data_to_delay(source_file, destination_file, mappings, dest_start_row):
    """
    Transfers data from a source file to a destination file based on column mappings provided.
    Skips rows where 'Observations' column contains 'cancel'.
    """
    # Load the workbooks and worksheets
    src_wb = load_workbook(source_file, read_only=True)
    src_ws = src_wb.active

    dest_wb = load_workbook(destination_file)
    dest_ws = dest_wb["Delays"]

    dest_ws.cell(row=1, column=1, value=destination_file.name[:-5])

    # # Print headers for debugging
    # print("Source headers:", [cell.value for cell in src_ws[3]])
    # print("Destination headers:", [cell.value for cell in dest_ws[2]])

    # Mapping source and destination headers to their respective column numbers
    src_header = {
        cell.value: col_num + 1
        for col_num, cell in enumerate(src_ws[3])
        if cell.value in mappings or any(cell.value in key for key in mappings if isinstance(key, tuple))
    }

    dest_header = {
        cell.value: col_num + 1
        for col_num, cell in enumerate(dest_ws[2])
        if cell.value in mappings.values()
    }

    # Find the column number for "Activity Summary" in the source file
    sum_col_num = None
    for col_num, cell in enumerate(src_ws[3]):
        if cell.value == "Activity Summary":
            sum_col_num = col_num + 1
            break

    if sum_col_num is None:
        print("Warning: 'Activity Summary' column not found in the source file.")
        return

    dest_row_counter = dest_start_row
    transferred_rows = 0

    # Iterating through each row in the source worksheet
    for row_num, row in enumerate(src_ws.iter_rows(min_row=4, values_only=True), 4):
        # Checks if the row is blank by looking at certain key columns
        key_column_indexes = [
            src_header['Planned Start'] - 1,
            src_header['Team Leader Name'] - 1,
            src_header['Date'] - 1
        ]
        if all(not row[idx] for idx in key_column_indexes):
            continue  # Skip the row as it is considered blank

        if sum_col_num and row[sum_col_num - 1] and "cancel" in row[sum_col_num - 1].lower():
            print(f"Skipping row {row_num} due to 'cancel' in Summary value : {row[sum_col_num - 1]}")
            continue

        # Transferring data based on mappings
        for src_col, dest_col in mappings.items():

            if src_col in src_header and dest_col in dest_header:
                dest_ws.cell(row=dest_row_counter, column=dest_header[dest_col]).value = row[
                    src_header[src_col] - 1]

            else:
                print(f"Missing columns '{src_col}' or '{dest_col}' in source or destination for row {row_num}.")

        dest_row_counter += 1
        transferred_rows += 1  # Increment only if data is actually transferred in this row

    # Save and close workbooks
    dest_wb.save(destination_file)
    src_wb.close()

    return transferred_rows


def calculate_week_num(date):
    # Adjust the date by moving any Sunday to the next day (Monday)
    adjusted_date = date + dt.timedelta(days=1) if date.weekday() == 6 else date

    # Use isocalendar to get the ISO week number
    _, iso_week, _ = adjusted_date.isocalendar()

    # Return the week number as a zero-padded string
    return f"{iso_week:02d}"



def derive_paths_from_date(dt_date):
    """
    Constructs various related paths based on a given date.
    """
    main_paths = define_related_paths()
    const_files_path = main_paths["Construction"]
    c_day, c_month, c_year = [dt_date.strftime(pattern) for pattern in ["%d", "%m", "%Y"]]
    c_week = calculate_week_num(dt_date)

    # Adjust the ISO year
    iso_year = dt_date.year
    if c_week == "01" and dt_date.month == 12:
        iso_year += 1

    c_formatted_dates = {
        "slash": f"{c_day}/{c_month}/{c_year[-2:]}",
        "dot": f"{c_day}.{c_month}.{c_year[-2:]}",
        "compact": f"{c_year[-2:]}{c_month}{c_day}",
    }

    paths = {
        "year": const_files_path / f"{iso_year}",
        "week": const_files_path / f"{iso_year}" / f"{c_week}",
        "day": const_files_path
               / f"{iso_year}"
               / f"{c_week}"
               / "Daily Reports"
    }

    # Return all required values
    return paths, c_formatted_dates, c_week



def derive_report_name(date, template="CIIM Report Table {}.xlsx"):

    return template.format(date)


def create_daily_report():
    # Prompt user to select a date
    str_date = cal_entry.entry.get()
    dt_date = datetime.strptime(str_date, '%Y-%m-%d')
    paths, c_formatted_dates, c_week = derive_paths_from_date(dt_date)


    # Ensure that  the Work plan was selected
    if len(cp_dates) == 0:
        error_message = "Please select the Construction plan and try again!"
        messagebox.showwarning("File Not Found", error_message)
        return

    # # Check if the week folder exists and create the next week folder if necessary
    # if c_week < 52:
    #     n_week_path = paths["n_week"]
    #     if not n_week_path.exists():
    #         try:
    #             n_week_path.mkdir(parents=True, exist_ok=True)
    #         except Exception as e:
    #             error_message = f"An error occurred while creating the directory: {e}"
    #             messagebox.showerror("Directory Creation Error", error_message)
    #     else:
    #         print(f"Directory already exists: {n_week_path}")
    # else:
    #     print(f"No need to create directory for week number {c_week} as it is not less than or equal to max week 52")

    # If the folder does not exist or the user chooses to overwrite, continue
    main_paths = define_related_paths()

    # Creating main paths
    for key in ["year", "week"]:
        Path(paths[key]).mkdir(parents=True, exist_ok=True)

    # Creating other necessary folders
    folders_to_create = [
        "Nominations",
        "Pictures",
        "Worklogs",
        "Toolboxes",
        "Daily Reports",
        "Weekly Reports",
        "Other",
    ]
    for folder in folders_to_create:
        (paths["week"] / folder).mkdir(exist_ok=True)

    # Derive report name and set paths
    ciim_daily_report = derive_report_name(c_formatted_dates["dot"])
    new_report_path = paths["day"] / ciim_daily_report

    # Check if the report file already exists
    if new_report_path.exists():
        overwrite = messagebox.askyesno("File Exists", f"{ciim_daily_report} already exists. Do you want to overwrite?")
        if not overwrite:
            return  # Exit the function if the user does not want to overwrite

    # Copy the template only if the report file does not already exist or if the user chose to overwrite
    if not new_report_path.exists():
        templates_path = main_paths["Templates"]
        fc_ciim_template_path = templates_path / DAILY_REPORT_TEMPLATE

        print(f'Copying template to: {new_report_path}')
        shutil.copy(fc_ciim_template_path, new_report_path)
    else:
        print(f"{new_report_path} already exists and was not copied.")

    # # Handle data report writing and copying
    # if Path(wp_path).parent != Path(paths["week"]):
    #     print("Not copying works to the selected date")
    #     return

    write_data_to_excel(
        wp_path,
        dt_date,
        c_formatted_dates["slash"],
        TO_DAILY_REPORT_MAPPINGS,
    )


def write_data_to_excel(src_path, dt_date, formatted_date, mappings, start_row=4):
    """
    Write data from the source Excel to a target report based on given mappings.
    """

    date_time = pd.to_datetime(formatted_date, format="%d/%m/%y", errors="coerce")
    report_path = extract_src_path_from_date(dt_date)
    report_name = report_path.name

    # Handle the case that the user chooses to overwrite and the file is open
    if is_file_locked(report_path):
        messagebox.showwarning(
            "File Locked",
            f"Please close {report_path.name} and try again."
        )
        print("Not overwriting since the file is open")
        return

    usecols_value = list(mappings.keys())
    df = pd.read_excel(src_path, skiprows=1, usecols=usecols_value)

    if df.empty:
        messagebox.showwarning("No Data", "No data found for the selected date.")
        return

    df["Date"] = pd.to_datetime(df["Date"], format="%d/%m/%Y", dayfirst=True, errors="coerce")
    target_df = df[df["Date"] == date_time]

    # Variables
    total_works_num = None
    planned_works_num = 0

    # Open the target workbook
    wb = load_workbook(filename=report_path)
    ws = wb.active

    try:
        # Insert the report filename in cell A1
        ws.cell(row=1, column=1, value=report_name[:-5])

        # Write headers (using the mappings values as headers)
        for col, header in enumerate(mappings.values(), 1):
            ws.cell(row=start_row - 1, column=col, value=header)

        # Write data
        for row_idx, (index, row_data) in enumerate(target_df.iterrows(), start=start_row):
            for col_idx, (src_header, dest_header) in enumerate(mappings.items(), 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data[src_header])
            total_works_num = row_idx

        # Delete irrelevant rows
        if total_works_num is not None:
            for row_idx in range(total_works_num + 1, ws.max_row + 1):
                ws.delete_rows(total_works_num + 1)

        # Input file
        # Iterate through  column 11 (Work Description) starting from row 4
        for row_idx in range(4, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=13).value

            if isinstance(cell_value, float):
                # Convert float value to a string1
                cell_value = str(cell_value)

            # Output file
            if not re.search(r"Cancel*", cell_value, re.IGNORECASE):
                # Replace the cell content with text
                ws.cell(
                    row=row_idx,
                    column=13,
                    value=""
                )
                planned_works_num += 1

        # Remove patterns from columns 10 (Foreman Name) and 11 (Team Leader Name)
        for row_idx in range(4, ws.max_row + 1):
            for col in [10, 11]:  # Check columns 10 and 11
                cell_value = ws.cell(row=row_idx, column=col).value
                if isinstance(cell_value, str):
                    # Remove the pattern "(...)" at the end
                    cleaned_value = re.sub(r'\s*\(.*?\)\s*$', '', cell_value)
                    ws.cell(row=row_idx, column=col, value=cleaned_value)

        cancelled_works_num = total_works_num - planned_works_num - 3  # (-3) because it starts from that row num
        messagebox.showinfo(title="Information",
                            message=f"Num of planned works (excluding cancelled): {planned_works_num}"
                                    f"\nNum of cancelled works: {cancelled_works_num}")

        wb.save(report_path)

    except ValueError as e:
        messagebox.showerror("Error", f"Failed to read Excel file: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

    finally:
        if 'wb' in locals():
            wb.close()  # Ensure the workbook is closed properly


def show_frame(frame_name):
    global current_frame

    for name, frame in frames.items():
        frame.pack_forget()
        if name == frame_name:
            frame.pack(fill="both", expand=True)
            current_frame = frame_name


def clock():
    time = datetime.now()
    hour = time.strftime(" %H:%M ")
    weekday = time.strftime("%A")
    day = time.strftime("%d")
    month = time.strftime("%m")
    year = time.strftime("%Y")

    hour_label.configure(text=hour)
    hour_label.after(6000, clock)

    day_label.configure(text=weekday + ", " + str(day) + "/" + str(month) + "/" + str(year))


def display_dist_list():
    global dist_list_populated

    show_frame("Dist list")
    paths = define_related_paths()
    # distlist_path = paths["Admin"] / "Distribution List.xlsx"

    distlist_path = "C:/Users/markpol/Grupo SEMI/CIIM - Admin Records/Distribution List.xlsx"


    if not dist_list_populated:
        try:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(distlist_path, usecols='A:B')

            # Iterate over the DataFrame and the text widgets at the same time
            for col, text_widget in zip(df.columns, text_widgets):
                # Clear the text widget first
                text_widget.delete('1.0', END)
                # Insert the data into the text widget
                column_data = '\n'.join(df[col].dropna().astype(str))
                text_widget.insert('1.0', column_data)
                # Highlight lines containing "cc" after inserting the text
                highlight_lines_containing_cc(text_widget)
                text_widget.config(state="disabled")

        except ValueError as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

        else:
            dist_list_populated = True


def display_phone_list():
    global is_phone_tree_populated

    # Ensure that  the Work plan was selected
    if len(cp_dates) == 0:
        error_message = "Please select the Construction plan and try again!"
        messagebox.showwarning("File Not Found", error_message)
        return

    show_frame("Phone")

    # Read the Excel file
    if not is_phone_tree_populated:
        try:
            phone_tree.heading("#0", text="Phone Numbers")
            df = pd.read_excel(wp_path, sheet_name="SEMI List", usecols="A, C, E, G")

            # Populate "Team Leaders" from DataFrame
            for department in organization["Team Leaders"]:
                if department in df:
                    workers = df[department].dropna().tolist()
                    workers = sorted(workers)  # sorts the name from A -> Z
                    organization["Team Leaders"][department].extend(workers)

            # Populate "Foremen" if the column exists in DataFrame
            if "Foremen" in df:
                foremen = df["Foremen"].dropna().tolist()
                foremen = sorted(foremen)  # sorts the name from A -> Z
                organization["Foremen"].extend(foremen)

            # Populate the tree with the organizational data
            for category, data in organization.items():
                category_id = phone_tree.insert('', 'end', text=category, open=False)
                if isinstance(data, list):
                    # If data is a list (like for "Foremen"), add items directly under the category
                    for item in data:
                        phone_tree.insert(category_id, 'end', text=item)
                else:
                    # If data is a dictionary (like for "Team Leaders"), iterate through departments
                    for dept, workers in data.items():
                        dept_id = phone_tree.insert(category_id, 'end', text=dept, open=False)
                        for worker in workers:
                            phone_tree.insert(dept_id, 'end', text=worker)

        except ValueError as e:
            messagebox.showerror("Error", f"Failed to read Excel file: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

        else:
            is_phone_tree_populated = True


# Function to toggle between template and original content
def dist_toggle_content(text_widget, template, original_contents, column):
    current_content = text_widget.get('1.0', 'end-1c')
    if current_content.strip() == template.strip():
        text_widget.config(state="normal")
        # If current content is the template, replace with original email
        text_widget.delete('1.0', 'end')
        text_widget.insert('1.0', original_contents[column])
        text_widget.config(state="disabled")
    else:
        # If current content is not the template, store it and insert template
        text_widget.config(state="normal")
        original_contents[column] = current_content
        text_widget.delete('1.0', 'end')
        text_widget.insert('1.0', template)
        text_widget.config(state="disabled")

    # Reapply the highlight to the text widget
    highlight_lines_containing_cc(text_widget)


def highlight_lines_containing_cc(text_widget):
    text_widget.tag_add("default_color", "1.0", "end")
    text_widget.tag_configure("highlight", underline=True, justify="center")
    text_widget.tag_configure("cc_highlight", font=("Roboto", 9, "bold"), background='#FB6C83')
    words_to_highlight = ["email", "whatsapp", "preview"]
    text_widget.tag_configure("email_color", foreground="#0489c9")

    # Searchs for lines containing '@' and apply the tag
    start_index = '1.0'
    while True:
        # Finds the next occurrence of '@'
        start_index = text_widget.search("@", start_index, 'end', nocase=True)
        if not start_index:
            break

        # Finds the end of the line containing '@'
        end_index = f"{start_index} lineend"
        # Applies the tag to the entire line
        text_widget.tag_add("email_color", f"{start_index} linestart", end_index)
        # Moves to the next line
        start_index = f"{end_index}+1c"

    # Insert new lines before "Preview Report (ISR)" if not already present
    start_index = '1.0'
    while True:
        start_index = text_widget.search("Preview Report (ISR)", start_index, 'end', nocase=True)
        if not start_index:
            break

        # Checks if the preceding characters are already '\n\n\n'
        if text_widget.get(f"{start_index}-3c", start_index) != '\n\n\n':
            text_widget.insert(start_index, '\n\n\n')
            start_index = f"{start_index}+{len('Preview Report (ISR)')}c"
        else:
            # Moves past the current match to continue searching
            start_index = f"{start_index}+{len('Preview Report (ISR)')}c"

    # Iterates over the list of words and highlight them with the 'highlight' tag
    for word in words_to_highlight:
        start_index = '1.0'
        while True:
            start_index = text_widget.search(word, start_index, 'end', nocase=True)
            if not start_index:
                break
            end_index = f"{start_index} lineend"
            text_widget.tag_add("highlight", start_index, end_index)
            start_index = f"{end_index}+1c"

    # Searchs and highlights 'cc' with a different tag 'cc_highlight'
    start_index = '1.0'
    while True:
        start_index = text_widget.search("cc", start_index, 'end', nocase=True)
        if not start_index:
            break
        end_index = f"{start_index} lineend"
        text_widget.tag_add("cc_highlight", start_index, end_index)
        start_index = f"{end_index}+1c"


def copy_to_clipboard(event):
    try:
        selected_item = phone_tree.selection()
        if selected_item:  # Check if something is selected
            parent = phone_tree.parent(selected_item[0])
            # Ensure the item is not a top-level category or department
            if parent and phone_tree.parent(parent):
                name = phone_tree.item(selected_item[0], 'text')
                frames["Phone"].clipboard_clear()
                frames["Phone"].clipboard_append(name)
                messagebox.showinfo("Info", f"Copied to clipboard: {name}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def open_procedure_file():
    # Ensure that  the Work plan was selected
    if len(cp_dates) == 0:
        error_message = "Please select the Construction plan and try again!"
        messagebox.showwarning("File Not Found", error_message)
        return

    paths = define_related_paths()
    proc_path = paths["procedure"]
    os.startfile(proc_path)


def open_wp_file():
    global wp_path
    os.startfile(wp_path)


def open_faults():
    paths = define_related_paths()
    faults_path = paths["Faults"]
    os.startfile(faults_path)


def delete_empty_rows(file_path, sheet_name, num_rows_to_keep):

    # Load the workbooks and worksheets
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Determine the total number of rows
    max_row = ws.max_row

    # Calculate the starting row to delete
    rows_to_delete_start = num_rows_to_keep + 1  # +1 to account for the first row to delete

    # Ensure we don't attempt to delete rows before the start of the worksheet
    if rows_to_delete_start <= max_row:
        # Delete rows from the end up to the calculated starting row
        for row in range(max_row, rows_to_delete_start - 1, -1):
            ws.delete_rows(row)

        print(f"{sheet_name}: Deleted rows after row {rows_to_delete_start}")

    wb.save(file_path)



def create_and_transfer_to_wkly_delay():
    """
    Transfers data to both the 'Cancellations' and 'Delays' sheets in the destination file.
    """
    global cp_dates

    try:
        # Ensure that  the Work plan was selected
        if len(cp_dates) == 0:
            error_message = "Please select the Construction plan and try again!"
            messagebox.showwarning("File Not Found", error_message)
            return

        # Variables for destination file paths and names
        delay_transferred_total = 0
        dest_start_row = 3
        main_paths = define_related_paths()
        date_str = cp_dates[0]
        formatted_str_date, dt_date, week_num = extract_date(date_str)
        delay_path = main_paths["Tracking"]

        # Set the delay filename and path
        wkly_delay_filename = f"Delays & Cancellations WW{week_num}.xlsx"
        new_report_path = delay_path / wkly_delay_filename
        templates_path = main_paths["Templates"]
        wkly_delay_temp_path = templates_path / WEEKLY_DELAY_TEMPLATE


        # Check if the file exists and copy template if necessary
        if not new_report_path.exists():
            shutil.copy(wkly_delay_temp_path, new_report_path)
            print(f'Copying template to: {new_report_path.name} and renaming to {new_report_path.name}')

        else:
            file_message_exist = f'{wkly_delay_filename[:-5]} already exists'
            overwrite = messagebox.askyesno("File Exists", f"{file_message_exist}\nDo you want to overwrite?")

            if not overwrite:
                return  # Exit the function if the user does not want to overwrite

            else:
                print(f'Overwriting {new_report_path}...')

        # Transfer data to Cancellations Sheet
        cancelled_transferred = transfer_data_to_cancelled(
            wp_path,
            new_report_path,
            TO_CANCELLED_MAPPING,
            dest_start_row
        )

        # Transfer data to Delays Sheet
        for date in cp_dates:
            formatted_str_date, dt_date, week_num = extract_date(date)
            daily_report_path = extract_src_path_from_date(dt_date)

            # Check if the daily report path exists
            if not os.path.exists(daily_report_path):
                error_message = f"File not found: {daily_report_path}. Stopping the transfer process."
                messagebox.showwarning("File Not Found", error_message)
                break

            delay_transferred = transfer_data_to_delay(
                daily_report_path,
                new_report_path,
                TO_DELAY_MAPPINGS,
                dest_start_row
            )
            dest_start_row += delay_transferred  # Update the starting row for the next date
            delay_transferred_total += delay_transferred

        # Message box with the results
        transferred_message = (
            f"{delay_transferred_total} rows were transferred to 'Delays' and "
            f"\n{cancelled_transferred} rows were transferred to 'Cancellations'"
        )
        messagebox.showinfo("Success", transferred_message)

        # Clean up empty rows
        delete_empty_rows(
            new_report_path,
            "Delays",
            delay_transferred_total + 2  # 2 accounted for the title and the Headers column
        )

        delete_empty_rows(
            new_report_path,
            "Cancellations",
            cancelled_transferred + 2  # 2 accounted for the title and the Headers column
        )

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def create_and_transfer_to_daily_delay(date_str):
    """
    Transfers data to both the 'Cancellations' and 'Delays' sheets in the destination file.
    """
    try:
        # Ensure that  the Work plan was selected
        if len(cp_dates) == 0:
            error_message = "Please select the Construction plan and try again!"
            messagebox.showwarning("File Not Found", error_message)
            return

        # Variables for destination file paths and names
        delay_transferred_total = 0
        dest_start_row = 3
        main_paths = define_related_paths()
        delay_path = main_paths["Tracking"]

        date = datetime.strptime(date_str, "%Y-%m-%d")
        # Now format the date as dd.mm.yy
        formatted_date = date.strftime("%d.%m.%y")  # Format the date to 'dd.mm.yy'

        # Set the delay filename and path
        daily_delay_filename = f"Delays & Cancellations {formatted_date}.xlsx"
        new_report_path = delay_path / daily_delay_filename
        templates_path = main_paths["Templates"]
        daily_delay_temp_path = templates_path / DAILY_DELAYS_CANC_TEMPLATE


        # Check if the file exists and copy template if necessary
        if not new_report_path.exists():
            shutil.copy(daily_delay_temp_path, new_report_path)
            print(f'Copying template to: {new_report_path.name} and renaming to {new_report_path.name}')

        else:
            file_message_exist = f'{daily_delay_filename[:-5]} already exists'
            overwrite = messagebox.askyesno("File Exists", f"{file_message_exist}\nDo you want to overwrite?")

            if not overwrite:
                return  # Exit the function if the user does not want to overwrite

            else:
                print(f'Overwriting {new_report_path}...')

        # Transfer data to Cancellations Sheet
        cancelled_transferred = transfer_data_to_cancelled(
            wp_path,
            new_report_path,
            TO_CANCELLED_MAPPING,
            dest_start_row
        )

        # Transfer data to Delays Sheet
        formatted_str_date, dt_date, week_num = extract_date(date_str)
        daily_report_path = extract_src_path_from_date(dt_date)

        # Check if the daily report path exists
        if not os.path.exists(daily_report_path):
            error_message = f"File not found: {daily_report_path}. Stopping the transfer process."
            messagebox.showwarning("File Not Found", error_message)
            return

        delay_transferred = transfer_data_to_delay(
            daily_report_path,
            new_report_path,
            TO_DELAY_MAPPINGS,
            dest_start_row
        )
        dest_start_row += delay_transferred  # Update the starting row for the next date
        delay_transferred_total += delay_transferred

        # Message box with the results
        transferred_message = (
            f"{delay_transferred_total} rows were transferred to 'Delays' and "
            f"\n{cancelled_transferred} rows were transferred to 'Cancellations'"
        )
        messagebox.showinfo("Success", transferred_message)

        # Clean up empty rows
        delete_empty_rows(
            new_report_path,
            "Delays",
            delay_transferred_total + 2  # 2 accounted for the title and the Headers column
        )

        delete_empty_rows(
            new_report_path,
            "Cancellations",
            cancelled_transferred + 2  # 2 accounted for the title and the Headers column
        )

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def create_and_transfer_to_wkly_ciim():

    # Ensure that  the Work plan was selected
    if len(cp_dates) == 0:
        error_message = "Please select the Construction plan and try again!"
        messagebox.showwarning("File Not Found", error_message)
        return

    # Process the first date
    first_date = cp_dates[0]
    formatted_date, dt_date, week_num = extract_date(first_date)
    wkly_ciim_folder_path = wp_path.parent / "Weekly Reports"
    wkly_ciim_filename = f"CIIM Report Table WW{week_num}.xlsx"
    new_report_path = wkly_ciim_folder_path / wkly_ciim_filename

    main_paths = define_related_paths()
    temp_path = main_paths["Templates"]
    daily_temp_path = temp_path / DAILY_REPORT_TEMPLATE

    # Check if the file exists and copy template if necessary
    if not new_report_path.exists():
        shutil.copy(str(daily_temp_path), str(new_report_path))
        print(f'Copied template to: {new_report_path.name}')

    # Check if the file is locked
    if is_file_locked(new_report_path):
        messagebox.showwarning(
            "File Locked",
            f"Please close {new_report_path.name} and try again."
        )
        return

    # Open the report file
    wb = load_workbook(filename=new_report_path, keep_links=True)
    ws = wb.active

    ws.cell(row=1, column=1, value=new_report_path.name[:-5])

    # Find the starting row for data in the destination file
    start_row = 4
    total_rows = 0  # Counter for total rows

    # Process the rest of the dates
    for date in cp_dates:
        formatted_date, dt_date, week_num = extract_date(date)
        daily_report_path = extract_src_path_from_date(dt_date)

        if not os.path.exists(daily_report_path):
            error_message = f"File not found: {daily_report_path}. Skipping this date."
            messagebox.showwarning("File Not Found", error_message)
            continue  # Skip this date and continue with the next

        # Read the data without skipping rows
        df = pd.read_excel(daily_report_path, skiprows=2)

        # Remove any empty rows in the DataFrame
        df.dropna(how='all', inplace=True)

        if df.empty:
            continue  # Skip if no data is available

        # Write the DataFrame to the destination file starting from the next available row
        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row):
                ws.cell(row=start_row + r_idx, column=c_idx + 1, value=value)

        # Update the total rows counter
        total_rows += len(df)

        # Update the starting row for the next batch of data
        start_row += len(df)

    # Remove empty rows from the destination sheet
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        if all(cell.value is None for cell in row):
            ws.delete_rows(row[0].row)

    # Show the total number of activities in a message box
    messagebox.showinfo("Activities Processed", f"Total number of activities: {total_rows}")

    # Save the workbook
    wb.save(new_report_path)
    print(f'Weekly CIIM Report saved to: {new_report_path}')

    ask_more = messagebox.askyesno("Continue?", f"Do you want to continue?")
    if not ask_more:
        return  # Exit the function if the user does not want to overwrite

    else:
        select_const_wp()
        create_and_transfer_to_wkly_ciim()


def update_menu_labels():
    """
    Update the menu labels based on the current value of ww_var.
    """
    global cp_dates

    if not cp_dates:
        print("No dates to update in the menu.")
        return

    # Remove existing date items before adding new ones
    create_menu.delete(0, END)  # Clears the menu

    week_num = ww_var.get()
    construction_wp = wp_path.name[:-5]
    # Print for debugging
    print(f"Updating Menu Labels with Week Number: {week_num}")

    create_menu.add_command(label=f"CIIM Report Table WW{week_num}", command=create_and_transfer_to_wkly_ciim)
    create_menu.add_command(label=f"Delays & Cancellations WW{week_num}", command=create_and_transfer_to_wkly_delay)

    for i, date in enumerate(cp_dates, start=2):  # Start the index from 2
        f_date = datetime.strptime(date, "%Y-%m-%d")  # Assuming the format is 'YYYY-MM-DD'
        # Now format the date as dd.mm.yy
        formatted_date = f_date.strftime("%d.%m.%y")  # Format the date to 'dd.mm.yy'

        # Use a default argument in the lambda to capture the current value of `date`
        create_menu.add_command(
            label=f"Delays & Cancellations {formatted_date}",
            command=lambda date=date: create_and_transfer_to_daily_delay(date)  # Capture date correctly
        )

    # Update Open Menu
    open_menu.entryconfig(0, label=construction_wp)



def naming_conversion():
    date = cp_dates[3]
    formatted_date, dt_date, week_num = extract_date(date)

    # Assuming paths and other data are correctly defined earlier in your code
    paths, c_formatted_dates, c_week = derive_paths_from_date(dt_date)

    if len(cp_dates) == 0:
        messagebox.showwarning("File Not Found", "Please select the Construction plan and try again!")
        return

    if not paths["week"].exists():
        messagebox.showwarning("Folder does not Exist",
                               f"{c_formatted_dates['compact']} folder does not exist. Please create the folder for that date first.")
        return

    folder_suffixes = {
        "Nominations": ["N1", "N2"]
    }

    # Regex to match file patterns
    pattern = re.compile(r'^(\d{1,2})_([A-Z]{2,})_(\d{4}-\d{2}-\d{2})_(N[1-2])\.(jpg|jpeg|pdf)$')

    def rename_file(file_path, folder_name):
        match = pattern.match(file_path.name)
        if not match:
            print(f"File {file_path.name} does not match the expected pattern.")
            return None

        ep_code, nm_code, date_part, suffix, extension = match.groups()
        new_name = f"{date_part}_{nm_code}_{ep_code}_{suffix}.{extension}"

        new_file_path = file_path.with_name(new_name)
        counter = 1

        # Ensure no overwrites
        while new_file_path.exists():
            new_name = f"{date_part}_{nm_code}_{ep_code}_{suffix}_{counter}.{extension}"
            new_file_path = file_path.with_name(new_name)
            counter += 1

        return new_file_path

    for folder in folder_suffixes.keys():
        folder_path = paths["week"] / folder
        if folder_path.exists() and folder_path.is_dir():
            for file in folder_path.iterdir():
                if file.is_file() and pattern.match(file.name):
                    print(f"Processing file: {file.name}")
                    new_file_path = rename_file(file, folder)
                    if new_file_path:
                        file.rename(new_file_path)
                        print(f"Renamed {file.name} to {new_file_path}")
                else:
                    print(f"File {file.name} skipped: does not match expected pattern.")



def delete_empty_folders():
    deleted_folders = []  # List to hold the paths of deleted folders
    # Prompt the user to select a folder
    folder_path = filedialog.askdirectory(title="Select a folder to delete empty subfolders")
    if not folder_path:
        return

    # Traverse the directory tree in reverse order
    for dirpath, dirnames, filenames in os.walk(folder_path, topdown=False):
        for dirname in dirnames:
            dir_to_check = os.path.join(dirpath, dirname)
            try:
                os.rmdir(dir_to_check)
                deleted_folders.append(dir_to_check)  # Add to the list of deleted folders
            except OSError:
                pass  # Ignore directories that couldn't be deleted

    # Create a message to display the result
    if deleted_folders:
        message = f"Deleted {len(deleted_folders)} empty folder(s)."
        messagebox.showinfo("Deletion Result", message)
    else:
        message = "No empty folders to delete."
        messagebox.showinfo("Deletion Result", message)


# ========================= Root config =========================
# Set DPI Awareness
enable_high_dpi_awareness()

app = ttk.Window(themename="darkly")
app.resizable(0, 0)
app.title("CIIM Manager")

# Grid
app.grid_columnconfigure(0, weight=1)
app.grid_rowconfigure(0, weight=1)
# Geometry
app_width = 750
app_height = 500
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
x = (screen_width / 2) - (app_width / 2)
y = (screen_height / 2) - app_height
app.geometry(f"{app_width}x{app_height}+{int(x)}+{int(y)}")

# ============================ Style ============================
style = ttk.Style()
style.configure("TButton", font=("Roboto", 9, "bold"), takefocus=False)
style.configure("TMenubutton", font=("Roboto", 9, "bold"))
style.configure("dark.Treeview.Heading", font=("Roboto", 9, "bold"), rowheight=40)
style.configure("dark.Treeview", rowheight=20, indent=50)

# =========================== Variables ===========================
current_frame = None

# Paths
base_path = Path("/")
wp_path = Path("/")
wp_var = StringVar()
current_combobox_index = StringVar()
# Lists and associated data
cp_dates = []
ww_var = IntVar()
# Miscellaneous variables
DAILY_REPORT_TEMPLATE = "CIIM Report Table - Template.xlsx"
WEEKLY_DELAY_TEMPLATE = "Delays & Cancellations - WEEKLY TEMPLATE.xlsx"
DAILY_DELAYS_CANC_TEMPLATE = "Delays & Cancellations - DAILY TEMPLATE.xlsx"

# List of Delay reasons
delay_reasons = [
    "Delay due to no TP",
    "Delay due to track vehicle maneuvers",
    "Delay due to waiting for the ISR/WSP Supervisor",
    "Delay due to waiting for the ISR Safety/ISR Comm. Supervisor",
    "Delay due to in the release of the electrified area/rail",
    "Delay due to coordination with the control center for the TP",
    "Delay due to track vehicle maneuvers",
    "Delay due to real hours are different for the 612",
    "--Other--"
]
# ============================ Mappings ============================
TO_DAILY_REPORT_MAPPINGS = {
    "Discipline": "Discipline",
    "Date": "Date",
    "T.P Start [Time]": "Planned Start",
    "T.P End [Time]": "Planned End",
    "EP": "EP",
    "ISR Start Section [Name]": "Start Section",
    "ISR  End Section [Name]": "End Section",
    "T.P Start [K.P]": "Start KP",
    "T.P End [K.P]": "End KP",
    "Foremen [Israel]": "Foreman Name",
    "Team Leader\nName (Phone)": "Team Leader Name",
    "Work Description": "Activity Description",
    "Observations": "Activity Summary"
}

TO_DELAY_MAPPINGS = {
    "Discipline": "Discipline",
    "Date": "Date",
    "Start Section":"Start Section",
    "End Section": "End Section",
    "Delay Cause": "Delay Cause",
    "Team Leader Name": "Team leader Name",
    "EP": "EP",
    "Activity Description": "Activity Description",
    "Toolbox": "Toolbox",
    "Worklog": "Worklog",
    "Planned Start": "Planned Start",
    "Actual Start": "Actual Start",
    "Planned End": "Planned End",
    "Actual End": "Actual End",
    "Mid Shift Delay": "Mid Shift Delay",
}

TO_CANCELLED_MAPPING = {
    "Date": "Date",
    "Discipline": "Discipline",
    "T.P Start [Time]": "Planned Start",
    "T.P End [Time]": "Planned End",
    "EP": "EP",
    "Team Leader\nName (Phone)": "Team leader Name",
    "Work Description": "Activity Description",
    "Observations": "Cancellation Cause",
}

# =========================== Frames ===========================
frames = {
    "Login": ttk.Frame(master=app),
    "Home": ttk.Frame(master=app),
    "Folder": ttk.Frame(master=app),
    "Phone": ttk.Frame(master=app),
    "Dist list": ttk.Frame(master=app),
    "Faults": ttk.Frame(master=app)
}

# ====================== Images ======================
images_dict = {
    "Home": 'images/home_l.png',
    "Folder": 'images/folder_l.png',
    "Phone": 'images/phone_l.png',
    "Dist list": 'images/mail_l.png',
    "Faults": 'images/faults_l.png',
    "Transfer": 'images/transfer_l.png',
}

photo_images = {}  # Dictionary to store the PhotoImage objects

# Convert each image path to a PhotoImage object and resize them
for key, path in images_dict.items():
    corrected_path = resource_path(path)  # Get the correct path
    image = Image.open(corrected_path)
    photo_image = ImageTk.PhotoImage(image)
    photo_images[key] = photo_image

# ====================== Home Side Frame ======================
side_frame = ttk.Frame(master=app, bootstyle="dark")
side_frame.pack(side=LEFT, fill=Y)

side_frame_empty = ttk.Label(master=side_frame, bootstyle="inverse.dark")
side_frame_empty.pack(fill='x', pady=40)

home_button = ttk.Button(master=side_frame, command=lambda: show_frame("Home"), bootstyle="dark",
                         image=photo_images["Home"], takefocus=False)
home_button.pack(fill='x', ipady=7)


folder_frame_button = ttk.Button(master=side_frame, command=lambda: show_frame("Folder"),
                         bootstyle="dark",
                         image=photo_images["Folder"], takefocus=False)
folder_frame_button.pack(fill='x', ipady=7)

distlist_button = ttk.Button(master=side_frame, command=lambda: display_dist_list(),
                             bootstyle="dark",
                             image=photo_images["Dist list"], takefocus=False)
distlist_button.pack(fill='x', ipady=7)

# phone_button = ttk.Button(master=side_frame, text="Phones", command=lambda: display_phone_list(),
#                           bootstyle="dark",
#                           image=photo_images["Phone"], takefocus=False)
# phone_button.pack(fill='x', ipady=7)

faults_button = ttk.Button(master=side_frame, text="Faults", command=lambda: show_frame("Faults"),
                          bootstyle="dark",
                          image=photo_images["Faults"], takefocus=False)
faults_button.pack(fill='x', ipady=7)

# ====================== Home Frame ======================

home_frame = frames["Home"]
home_frame.columnconfigure(0, weight=0)
home_frame.columnconfigure(1, weight=1)
home_frame.rowconfigure(1, weight=1)

# Top Frame
top_frame = ttk.Frame(master=home_frame)
top_frame.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)

# Packing the hour and day labels at the top first
hour_label = ttk.Label(master=top_frame, text="12:49", font="digital-7 120")
hour_label.pack(anchor="center")

day_label = ttk.Label(master=top_frame, text="Saturday 22/01/2023", font="digital-7 35", style="secondary")
day_label.pack(padx=5, pady=5)

# Bottom Frame
bottom_frame = ttk.Frame(master=home_frame)
bottom_frame.grid(row=2, column=1, sticky='nsew', padx=5, pady=5)

home_browse_button = ttk.Button(master=bottom_frame, text="Select", command=select_const_wp, width=8,
                                takefocus=False, bootstyle="secondary")
home_browse_button.pack(anchor='sw', side='left', pady=5)
path_entry = ttk.Entry(master=bottom_frame, textvariable=wp_var, bootstyle="danger")
path_entry.pack(anchor='s', side='left', fill='x', expand=True, pady=5)


# Menus

# Open files menu
open_mb = ttk.Menubutton(top_frame, text="Open", width=8)
open_mb.pack(pady=30)

open_menu = ttk.Menu(open_mb, tearoff=0)

open_menu.add_command(label="Construction Work Plan", command=open_wp_file)
open_menu.add_command(label="Fault Report Database", command=open_faults)
open_menu.add_command(label="Procedure", command=open_procedure_file)

open_mb["menu"] = open_menu



# Create files menu
create_mb = ttk.Menubutton(top_frame, text="Create", width=8, bootstyle = "Success")
create_mb.pack()

create_menu = ttk.Menu(open_mb, tearoff=0)

create_mb["menu"] = create_menu

del_button = ttk.Button(master=top_frame, text="Delete Empty", command=delete_empty_folders, width=15,
                                takefocus=False, bootstyle="secondary")
del_button.pack(anchor="sw")


# ====================== Phones frame ======================
is_phone_tree_populated = False  # it will ensure it runs only once and not each time we launch the frame

phones_frame = frames["Phone"]
phones_frame.pack(fill="both", expand=True)

phones_frame.rowconfigure(0, weight=1)
phones_frame.columnconfigure(0, weight=1)

phone_tree_scroll = ttk.Scrollbar(phones_frame, style="round")
phone_tree_scroll.grid(row=0, column=1, sticky="nsw")

phone_tree = ttk.Treeview(phones_frame, cursor="hand2", yscrollcommand=phone_tree_scroll.set,
                          style="dark.Treeview", padding=10)
phone_tree.grid(row=0, column=0, sticky="nsew")

organization = {
    "Team Leaders": {
        "OCS": [],
        "SCADA": [],
        "SURICATA": [],
        "OCS-D": [],
    },
    "Foremen": []
}

phone_tree_scroll.config(command=phone_tree.yview)
# Bindings
phone_tree.bind('<Double-1>', copy_to_clipboard)

# ====================== Dist. list frame ======================
def get_dates():
    today = dt.date.today()
    tonight = today
    tomorrow = today + dt.timedelta(days=1)
    return tonight.strftime('%d.%m.%y'), tomorrow.strftime('%d.%m.%y')


def fill_template(template, tonight_date, tomorrow_date):
    # Replace all occurrences of dd.mm.yy with the correct dates
    return template.replace("dd.mm.yy", tonight_date, 1).replace("dd.mm.yy", tomorrow_date, 1).replace("dd.mm.yy", tonight_date, 1).replace("dd.mm.yy", tomorrow_date, 1)


def populate_templates_with_dates(templates):
    tonight_date, tomorrow_date = get_dates()
    for key, template in templates.items():
        templates[key] = fill_template(template, tonight_date, tomorrow_date)
    return templates

# Initialize frame configuration
dist_list_populated = False  # it will ensure it runs only once and not each time we launch the frame
dist_frame = frames["Dist list"]

# Configure the frame to give equal weight to all columns
for i in range(3):
    dist_frame.columnconfigure(i, weight=1)
    dist_frame.rowconfigure(1, weight=1)

templates = {
    "Preview":
               "Hi all,"
               "\n\nAttached is the CIIM preview report."
    "\n\nPlease note that the attached is a draft and may contain missing information or require further updates or clarification.",
    "Work plan": "                 Email:"
                    "\n\nHi all,"
                    "\n\nAttached is the updated work plan for tonight (dd.mm.yy) and tomorrow morning (dd.mm.yy)."
                    "\n\n\n              Whatsapp Message"
                    "\n\nGood afternoon,\nAttached is the work plan for tonight (dd.mm.yy) and "
                    "tomorrow morning (dd.mm.yy)."
                    "\nPlease note that the hours listed are the starting hours of the TP."
                    "\nPlease keep in touch with your managers about the time you should be in the field."
                    "\nGood luck.",
}

templates = populate_templates_with_dates(templates)

# Stores the original content of the text widgets
original_contents = ['' for _ in range(4)]

# Text widgets list
text_widgets = [Text(dist_frame) for _ in range(4)]


def make_command(col, tw, temp):
    return lambda: dist_toggle_content(tw, temp, original_contents, col)


# Creates buttons and text widgets, and place them in the frame
for column, (label_text, template) in enumerate(templates.items()):
    button = ttk.Button(dist_frame, text=label_text, command=make_command(column, text_widgets[column], template),
                        bootstyle="link", takefocus=False)
    button.grid(row=0, column=column, pady=5)
    text_widget = text_widgets[column]
    # text_widget.config(highlightbackground="#d3d3d3")
    text_widget.grid(row=1, column=column, sticky="nsew", padx=2)
    ToolTip(button, text="Click for template/emails", delay=500)


# ====================== Faults Frame ======================

# Function to generate the email content
def generate_faults_email():
    # Get fault number and department selection
    fault_number = fault_number_entry.get()
    department = department_var.get()

    # Validate fault number (must be 7 digits)
    if len(fault_number) != 7 or not fault_number.isdigit():
        messagebox.showerror("Invalid Input", "Fault number must be 7 digits.")
        return

    # Validate department selection
    if department not in department_options:
        messagebox.showerror("Invalid Input", "Please select a valid department.")
        return

    # Prepare the email content based on department
    confirmation_mail = "We confirm the reception of the Fault Report No. {}.\n".format(fault_number)

    email_content = ""

    if department == "OCS":
        email_content += "ja.sierra@syneox.com\naturk@gruposemi.com\n\n"
        email_content += "OCS Fault report No. {}\n\n".format(fault_number)
        email_content += "Hi All,\n\nAttached is the Fault Report No. {}.\n".format(fault_number)

    elif department == "SCADA":
        email_content += "yshoshany@gruposemi.com\nmmiran@gruposemi.com\ngmaskalchi@gruposemi.com\n\n"
        email_content += "SCADA Fault report No. {}\n\n".format(fault_number)
        email_content += "Hi All,\n\nAttached is the Fault Report No. {}.\n".format(fault_number)

    elif department == "TS":
        email_content += "aturk@gruposemi.com\nCC:\narodriguez@gruposemi.com\nygutmacher@gruposemi.com\n\n"
        email_content += "TS Fault report No. {}\n\n".format(fault_number)
        email_content += "Hi Ali,\n\nAttached is the Fault Report No. {}.\n".format(fault_number)

    # Update the text widgets with the generated email content
    confirmation_text_widget.delete(1.0, END)
    confirmation_text_widget.insert(END, confirmation_mail)

    email_text_widget.delete(1.0, END)
    email_text_widget.insert(END, email_content)

    # Apply highlighting to the text widgets
    apply_highlighting(confirmation_text_widget, fault_number)
    apply_highlighting(email_text_widget, fault_number)

# Function to highlight emails, 'cc', and specific fault report lines in a text widget
def apply_highlighting(text_widget, fault_number):
    # Configure text tags for highlighting
    text_widget.tag_configure("email_color", foreground="#0489c9")
    text_widget.tag_configure("cc_highlight", font=("Roboto", 9, "bold"), background='#FB6C83')
    text_widget.tag_configure("bold_text", font=("Roboto", 9, "bold"))

    # Highlight email addresses (containing '@')
    start_index = '1.0'
    while True:
        start_index = text_widget.search("@", start_index, 'end', nocase=True)
        if not start_index:
            break
        end_index = f"{start_index} lineend"
        text_widget.tag_add("email_color", f"{start_index} linestart", end_index)
        start_index = f"{end_index}+1c"

    # Highlight 'cc' occurrences
    start_index = '1.0'
    while True:
        start_index = text_widget.search("CC:", start_index, 'end', nocase=True)
        if not start_index:
            break
        end_index = f"{start_index} lineend"
        text_widget.tag_add("cc_highlight", start_index, end_index)
        start_index = f"{end_index}+1c"

    # Highlight specific fault report lines
    patterns = [
        f"OCS Fault report No. {fault_number}",
        f"SCADA Fault report No. {fault_number}",
        f"TS Fault report No. {fault_number}"
    ]

    for pattern in patterns:
        start_index = '1.0'
        while True:
            start_index = text_widget.search(pattern, start_index, 'end', nocase=True)
            if not start_index:
                break
            end_index = f"{start_index}+{len(pattern)}c"
            text_widget.tag_add("bold_text", start_index, end_index)
            start_index = end_index

# Create a frame for the content
faults_frame = frames["Faults"]
faults_frame.pack(pady=20, padx=20, fill="both", expand=True)

# Configure grid for layout
faults_frame.columnconfigure(1, weight=1)

# Entry widget for fault number
fault_number_label = ttk.Label(faults_frame, text="Fault Number (7 digits):")
fault_number_label.grid(row=1, column=0, sticky="w", pady=5, padx=5)

fault_number_entry = ttk.Entry(faults_frame, width=12)
fault_number_entry.grid(row=1, column=1, sticky="w", pady=5, padx=5)

# Dropdown menu for department selection
department_label = ttk.Label(faults_frame, text="Select Department:")
department_label.grid(row=2, column=0, sticky="w", pady=5, padx=5)

department_options = ["OCS", "SCADA", "TS"]
department_var = StringVar()
department_dropdown = ttk.Combobox(faults_frame, textvariable=department_var, values=department_options,
                                   state="readonly", width=10)
department_dropdown.grid(row=2, column=1, sticky="w", pady=5, padx=5)

# Generate button
generate_button = ttk.Button(faults_frame, text="Generate", command=generate_faults_email)
generate_button.grid(row=3, column=0, columnspan=2, pady=20, padx=5)

# Label and text widget for confirmation mail
confirmation_label = ttk.Label(faults_frame, text="2. Click 'Reply' with the following:")
confirmation_label.grid(row=4, column=0, columnspan=2, sticky="w", pady=5, padx=5)

confirmation_text_widget = Text(faults_frame, height=3, width=70)
confirmation_text_widget.grid(row=5, column=0, columnspan=2, sticky="ew", pady=5, padx=5)

# Label and text widget for department-specific email
email_label = ttk.Label(faults_frame, text="3. Send an email with the Fault Report attached and write the following:")
email_label.grid(row=6, column=0, columnspan=2, sticky="w", pady=5, padx=5)

email_text_widget = Text(faults_frame, height=13, width=70)
email_text_widget.grid(row=7, column=0, columnspan=2, sticky="ew", pady=5, padx=5)

# ====================== Folder Frame ======================

folder_frame = frames["Folder"]

folder_frame.rowconfigure(0, weight=1)
folder_frame.columnconfigure(0, weight=1)


folder_frame_top = ttk.Frame(master=folder_frame)
folder_frame_top.grid(row=0, column=0, sticky='nsew')

select_folder_label = ttk.Label(master=folder_frame_top, text="   Select date:  ")
select_folder_label.grid(row=0, column=0, padx=5, pady=43, sticky="e")

cal_entry = ttk.DateEntry(folder_frame_top, bootstyle="danger", dateformat="%Y-%m-%d")
cal_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")


folder_frame_toolbar = ttk.Frame(master=folder_frame)
folder_frame_toolbar.grid(row=1, columnspan=3, sticky="nsew")

tab2_seperator = ttk.Separator(folder_frame_toolbar, orient="horizontal")
tab2_seperator.pack(side=TOP, fill=BOTH)

# Button
create_button = ttk.Button(master=folder_frame_toolbar, text="Create", command=create_daily_report, width=10)
create_button.pack(side=RIGHT, padx=10, pady=10)

nc_button = ttk.Button(master=folder_frame_toolbar, text="Naming Convension", command=naming_conversion, width=17, style='secondary')
nc_button.pack(side=LEFT, padx=10, pady=10)


show_frame("Home")
clock()
app.mainloop()
